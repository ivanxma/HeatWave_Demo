import base64
import csv
import hashlib
import importlib
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler
import io
from itertools import zip_longest
import json
import os
import re
import secrets
import socket
import ssl
import subprocess
import sys
import threading
import time
import urllib.request
from functools import wraps
from pathlib import Path

import mysql.connector
from openpyxl import load_workbook
from flask import Flask, Response, flash, jsonify, redirect, render_template, request, session, url_for
from mysql.connector import errorcode
from werkzeug.serving import ThreadedWSGIServer, WSGIRequestHandler, is_ssl_error, load_ssl_context


APP_TITLE = "HeatWave Demo"
ROOT_DIR = Path(__file__).resolve().parent
PROFILE_STORE = ROOT_DIR / "profiles.json"
APP_SLUG = "heatwave-demo"
UPDATE_STATUS_FILE = Path(os.environ.get("HEATWAVE_DEMO_UPDATE_STATUS", f"/tmp/{APP_SLUG}-update-status.json"))
UPDATE_LOG_FILE = Path(os.environ.get("HEATWAVE_DEMO_UPDATE_LOG", f"/tmp/{APP_SLUG}-update.log"))
UPDATE_WORKER_FILE = ROOT_DIR / "heatwave_demo_update_worker.py"
VERSION_FILE = ROOT_DIR / "appver.json"
VERSION_CHECK_TIMEOUT_SECONDS = 5
VERSION_CHECK_GIT_TIMEOUT_SECONDS = 10
IMPORT_PREVIEW_DIR = Path("/tmp/hw_nlsql_imports")
DEFAULT_PROFILE = {
    "name": "",
    "host": "",
    "port": 3306,
    "database": "performance_schema",
    "ssh_enabled": False,
    "ssh_host": "",
    "ssh_port": 22,
    "ssh_user": "",
    "ssh_key_file": "",
    "connection_timeout": "",
    "read_timeout": "",
    "write_timeout": "",
    "max_execution_time": "",
    "wait_timeout": "",
    "interactive_timeout": "",
}
LOGIN_VALIDATION_CONNECTION_TIMEOUT_SECONDS = 5
SESSION_VALIDATION_CONNECTION_TIMEOUT_SECONDS = 3
DEFAULT_CONNECTION_TIMEOUT_SECONDS = 5
DEFAULT_TLS_HANDSHAKE_TIMEOUT_SECONDS = 5
DEFAULT_REQUEST_SOCKET_TIMEOUT_SECONDS = 30
CONNECTION_CACHE_MAX_AGE_SECONDS = 3600
SYSTEM_DATABASES = {"information_schema", "mysql", "performance_schema", "sys"}
DASHBOARD_TABS = [
    {"id": "demo", "label": "Demo"},
    {"id": "server-info", "label": "Server Info"},
]
DB_ADMIN_TABS = [
    {"id": "db", "label": "DB"},
    {"id": "table", "label": "Table"},
    {"id": "hw-tables", "label": "HW Tables"},
    {"id": "monitoring", "label": "Monitoring"},
]
DB_ADMIN_MONITORING_VIEWS = {
    "heatwave-performance-query",
    "heatwave-ml-query",
    "hw-table-load-recovery",
}
MYSQL_TYPE_OPTIONS = [
    {"name": "INT", "label": "INT", "param_mode": "none", "param_placeholder": ""},
    {"name": "BIGINT", "label": "BIGINT", "param_mode": "none", "param_placeholder": ""},
    {"name": "VARCHAR", "label": "VARCHAR", "param_mode": "single", "param_placeholder": "255"},
    {"name": "CHAR", "label": "CHAR", "param_mode": "single", "param_placeholder": "36"},
    {"name": "DECIMAL", "label": "DECIMAL", "param_mode": "multi", "param_placeholder": "10,3"},
    {"name": "DATE", "label": "DATE", "param_mode": "none", "param_placeholder": ""},
    {"name": "DATETIME", "label": "DATETIME", "param_mode": "none", "param_placeholder": ""},
    {"name": "TIMESTAMP", "label": "TIMESTAMP", "param_mode": "none", "param_placeholder": ""},
    {"name": "TEXT", "label": "TEXT", "param_mode": "none", "param_placeholder": ""},
    {"name": "LONGTEXT", "label": "LONGTEXT", "param_mode": "none", "param_placeholder": ""},
    {"name": "JSON", "label": "JSON", "param_mode": "none", "param_placeholder": ""},
    {"name": "BOOLEAN", "label": "BOOLEAN", "param_mode": "none", "param_placeholder": ""},
]
MYSQL_TYPE_OPTION_MAP = {row["name"]: row for row in MYSQL_TYPE_OPTIONS}
IDENTIFIER_RE = re.compile(r"^[A-Za-z0-9_$]+$")
MYSQL_DATA_TYPE_RE = re.compile(r"^[A-Z]+(?:\s+[A-Z]+)?(?:\(\s*\d+(?:\s*,\s*\d+)?\s*\))?$")
ASKME_SCHEMA_NAME = "askme"
ASKME_CONFIG_ROWS = [
    ("OCI_REGION", ""),
    ("OCI_CONFIG_FILE", "~/.oci/config"),
    ("OCI_CONFIG_PROFILE", "DEFAULT"),
    ("OCI_BUCKET_NAME", ""),
    ("OCI_NAMESPACE", ""),
    ("OCI_BUCKET_FOLDER", "askme_user_data/user_documents-"),
]
NAV_GROUPS = [
    {
        "label": "Home",
        "items": [{"endpoint": "home", "label": "Dashboard"}],
    },
    {
        "label": "Admin",
        "items": [
            {"endpoint": "connection_profile", "label": "Connection Profile"},
            {"endpoint": "db_admin_page", "label": "DB Admin"},
            {"endpoint": "import_page", "label": "Import"},
            {"endpoint": "setup_configdb_page", "label": "Setup configdb"},
            {"endpoint": "setup_askme_page", "label": "Setup ObjectStorage"},
            {"endpoint": "update_heatwave_demo", "label": "Update HeatWave_Demo"},
        ],
    },
    {
        "label": "HeatWave",
        "items": [
            {"endpoint": "nlsql_page", "label": "NL_SQL"},
            {"endpoint": "vision_page", "label": "HWVision"},
            {"endpoint": "heatwave_genai_page", "label": "GenAI"},
            {"endpoint": "askme_genai_page", "label": "Askme GenAI"},
            {"endpoint": "heatwave_ml_page", "label": "HeatWave ML"},
            {"endpoint": "heatwave_lh_external_page", "label": "HeatWave LH/External Table"},
        ],
    },
]
HEATWAVE_PERFORMANCE_SQL = {
    "innodb": """
SELECT /*+ SET_VAR(use_secondary_engine=off) */ airline.airlinename, SUM(booking.price) as price_tickets, count(*) as nb_tickets
      FROM airportdb.booking booking, airportdb.flight flight, airportdb.airline airline, airportdb.airport_geo airport_geo
      WHERE booking.flight_id=flight.flight_id
      AND airline.airline_id=flight.airline_id
      AND flight.from=airport_geo.airport_id
      AND airport_geo.country = "UNITED STATES"
      GROUP BY airline.airlinename
      ORDER BY nb_tickets desc, airline.airlinename
      LIMIT 10;
""".strip(),
    "rapid": """
SELECT /*+ SET_VAR(use_secondary_engine=on) */ airline.airlinename, SUM(booking.price) as price_tickets, count(*) as nb_tickets
      FROM airportdb.booking booking, airportdb.flight flight, airportdb.airline airline, airportdb.airport_geo airport_geo
      WHERE booking.flight_id=flight.flight_id
      AND airline.airline_id=flight.airline_id
      AND flight.from=airport_geo.airport_id
      AND airport_geo.country = "UNITED STATES"
      GROUP BY airline.airlinename
      ORDER BY nb_tickets desc, airline.airlinename
      LIMIT 10;
""".strip(),
}
HEATWAVE_PERFORMANCE_EXEC_SQL = {
    "innodb": """
SELECT /*+ SET_VAR(use_secondary_engine=off) */ airline.airlinename, SUM(booking.price) as price_tickets, count(*) as nb_tickets
      FROM airportdb.booking booking, airportdb.flight flight, airportdb.airline airline, airportdb.airport_geo airport_geo
      WHERE booking.flight_id=flight.flight_id
      AND airline.airline_id=flight.airline_id
      AND flight.from=airport_geo.airport_id
      AND airport_geo.country = "UNITED STATES"
      GROUP BY airline.airlinename
      ORDER BY nb_tickets desc, airline.airlinename
      LIMIT 10
""".strip(),
    "rapid": """
SELECT /*+ SET_VAR(use_secondary_engine=on) */ airline.airlinename, SUM(booking.price) as price_tickets, count(*) as nb_tickets
      FROM airportdb.booking booking, airportdb.flight flight, airportdb.airline airline, airportdb.airport_geo airport_geo
      WHERE booking.flight_id=flight.flight_id
      AND airline.airline_id=flight.airline_id
      AND flight.from=airport_geo.airport_id
      AND airport_geo.country = "UNITED STATES"
      GROUP BY airline.airlinename
      ORDER BY nb_tickets desc, airline.airlinename
      LIMIT 10
""".strip(),
}
PREFERRED_DEFAULT_MODEL = "meta.llama-3.3-70b-instruct"
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("FLASK_SECRET_KEY", "change-this-secret-key")
app.config["SESSION_COOKIE_NAME"] = "heatwave_demo_session"
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
sys.modules.setdefault("app_context", sys.modules[__name__])
_CONNECTION_CACHE = {}
_CONNECTION_CACHE_LOCK = threading.RLock()


def ensure_profile_store():
    if PROFILE_STORE.exists():
        return
    PROFILE_STORE.write_text(json.dumps({"profiles": []}, indent=2), encoding="utf-8")


def _normalized_port(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return DEFAULT_PROFILE["port"]


def _normalized_optional_timeout(value, *, allow_zero=False):
    if value in (None, ""):
        return ""
    try:
        normalized = int(value)
    except (TypeError, ValueError):
        return ""
    if normalized > 0:
        return normalized
    if allow_zero and normalized == 0:
        return 0
    return ""


def normalize_profile(payload):
    ssh_enabled_value = payload.get("ssh_enabled", False)
    ssh_enabled = str(ssh_enabled_value).strip().lower() in {"1", "true", "yes", "on"}
    return {
        "name": str(payload.get("name", "")).strip(),
        "host": str(payload.get("host", "")).strip(),
        "port": _normalized_port(payload.get("port")),
        "database": str(payload.get("database", "")).strip(),
        "ssh_enabled": ssh_enabled,
        "ssh_host": str(payload.get("ssh_host", "")).strip(),
        "ssh_port": _normalized_port(payload.get("ssh_port") or DEFAULT_PROFILE["ssh_port"]),
        "ssh_user": str(payload.get("ssh_user", "")).strip(),
        "ssh_key_file": str(payload.get("ssh_key_file", "")).strip(),
        "connection_timeout": _normalized_optional_timeout(payload.get("connection_timeout")),
        "read_timeout": _normalized_optional_timeout(payload.get("read_timeout")),
        "write_timeout": _normalized_optional_timeout(payload.get("write_timeout")),
        "max_execution_time": _normalized_optional_timeout(payload.get("max_execution_time"), allow_zero=True),
        "wait_timeout": _normalized_optional_timeout(payload.get("wait_timeout")),
        "interactive_timeout": _normalized_optional_timeout(payload.get("interactive_timeout")),
    }


def load_profiles():
    ensure_profile_store()
    try:
        data = json.loads(PROFILE_STORE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    rows = data.get("profiles", [])
    profiles = []
    for row in rows:
        profile = normalize_profile(row)
        if profile["name"]:
            profiles.append(profile)
    return sorted(profiles, key=lambda item: item["name"].lower())


def save_profiles(profiles):
    normalized = []
    seen_names = set()
    for row in profiles:
        profile = normalize_profile(row)
        if not profile["name"]:
            continue
        key = profile["name"].lower()
        if key in seen_names:
            continue
        seen_names.add(key)
        normalized.append(profile)
    PROFILE_STORE.write_text(
        json.dumps({"profiles": normalized}, indent=2),
        encoding="utf-8",
    )


def validate_profile_settings(profile, *, require_key_file=False):
    errors = []
    if not profile.get("host"):
        errors.append("Host is required.")
    if not profile.get("port"):
        errors.append("Port is required.")
    if profile.get("ssh_enabled"):
        if not profile.get("ssh_host"):
            errors.append("SSH jump host is required when SSH tunnel is enabled.")
        if not profile.get("ssh_user"):
            errors.append("SSH user is required when SSH tunnel is enabled.")
        if not profile.get("ssh_port"):
            errors.append("SSH port is required when SSH tunnel is enabled.")
        key_file = str(profile.get("ssh_key_file", "")).strip()
        if not key_file:
            errors.append("SSH private key path is required when SSH tunnel is enabled.")
        elif require_key_file and not Path(os.path.expanduser(key_file)).exists():
            errors.append("SSH private key file does not exist: {}".format(key_file))
    return errors


def get_profile_by_name(profile_name):
    lookup = str(profile_name or "").strip().lower()
    for profile in load_profiles():
        if profile["name"].lower() == lookup:
            return profile
    return None


def get_session_profile():
    payload = session.get("connection_profile", {})
    if not payload:
        return normalize_profile(DEFAULT_PROFILE)
    return normalize_profile(payload)


def set_session_profile(profile):
    new_profile = normalize_profile(profile)
    old_profile = get_session_profile()
    if old_profile != new_profile:
        clear_login_state(keep_profile=False)
    session["connection_profile"] = new_profile
    session["profile_name"] = new_profile["name"]


def get_selected_profile_name():
    requested = request.args.get("profile", "").strip()
    if requested:
        return requested
    requested = request.values.get("profile_name", "").strip()
    if requested:
        return requested
    return str(session.get("profile_name", "")).strip()


def profile_is_complete(profile=None):
    active = profile or get_session_profile()
    return bool(active["host"] and active["database"] and active["port"])


def get_connection_summary():
    profile = get_session_profile()
    if not profile_is_complete(profile):
        return "Not configured"
    summary = "{host}:{port}/{database}".format(**profile)
    if profile.get("ssh_enabled"):
        summary = "{} via SSH {}@{}:{}".format(
            summary,
            profile.get("ssh_user") or "-",
            profile.get("ssh_host") or "-",
            profile.get("ssh_port") or 22,
        )
    return summary


def fetch_connection_timeout_settings():
    default_values = {
        "connection_timeout": "",
        "read_timeout": "",
        "write_timeout": "",
        "max_execution_time": "",
        "wait_timeout": "",
        "interactive_timeout": "",
    }
    profile = get_session_profile()
    default_values["connection_timeout"] = profile.get("connection_timeout") or DEFAULT_CONNECTION_TIMEOUT_SECONDS
    if not session.get("logged_in"):
        return default_values
    try:
        rows = run_sql(
            """
            select
              @@session.net_read_timeout as net_read_timeout,
              @@session.net_write_timeout as net_write_timeout,
              @@session.max_execution_time as max_execution_time,
              @@session.wait_timeout as wait_timeout,
              @@session.interactive_timeout as interactive_timeout
            """,
            include_database=False,
        )
    except mysql.connector.Error:
        return default_values
    row = rows[0] if rows else ("", "", "", "", "")
    default_values["read_timeout"] = row[0]
    default_values["write_timeout"] = row[1]
    default_values["max_execution_time"] = row[2]
    default_values["wait_timeout"] = row[3]
    default_values["interactive_timeout"] = row[4]
    return default_values


def get_connection_timeout_summary(timeout_settings=None):
    values = timeout_settings or fetch_connection_timeout_settings()
    if not any(str(values.get(key, "")).strip() for key in values):
        return "Unavailable"
    return "/".join(
        (
            str(values["connection_timeout"] or "-"),
            str(values["read_timeout"] or "-"),
            str(values["write_timeout"] or "-"),
            str(values["max_execution_time"] if values["max_execution_time"] != "" else "-"),
        )
    )


def clear_login_state(keep_profile=True):
    clear_cached_mysql_connections_for_session()
    profile = get_session_profile() if keep_profile else None
    session["logged_in"] = False
    session["db_user"] = ""
    session["db_password"] = ""
    session.pop("connection_cache_id", None)
    session.pop("llm_model_cache", None)
    if keep_profile and profile:
        session["connection_profile"] = profile
        session["profile_name"] = profile["name"]
    elif not keep_profile:
        session.pop("connection_profile", None)
        session.pop("profile_name", None)


def start_connection_cache_session():
    clear_cached_mysql_connections_for_session()
    session["connection_cache_id"] = secrets.token_urlsafe(18)


def get_connection_config(
    user=None,
    password=None,
    include_database=True,
    *,
    fallback_connection_timeout=None,
    fallback_read_timeout=None,
    fallback_write_timeout=None,
):
    profile = get_session_profile()
    config = {
        "host": profile["host"],
        "port": profile["port"],
    }
    if include_database and profile["database"]:
        config["database"] = profile["database"]
    if profile["connection_timeout"]:
        config["connection_timeout"] = profile["connection_timeout"]
    elif fallback_connection_timeout is not None:
        config["connection_timeout"] = int(fallback_connection_timeout)
    if profile["read_timeout"]:
        config["read_timeout"] = profile["read_timeout"]
    elif fallback_read_timeout is not None:
        config["read_timeout"] = int(fallback_read_timeout)
    if profile["write_timeout"]:
        config["write_timeout"] = profile["write_timeout"]
    elif fallback_write_timeout is not None:
        config["write_timeout"] = int(fallback_write_timeout)
    resolved_user = session.get("db_user", "") if user is None else user
    resolved_password = session.get("db_password", "") if password is None else password
    if resolved_user:
        config["user"] = resolved_user
    if resolved_password:
        config["password"] = resolved_password
    return config


def _apply_connection_profile_session_settings(cnx, profile=None):
    active_profile = normalize_profile(profile or get_session_profile())
    assignments = []
    if active_profile["read_timeout"]:
        assignments.append(("net_read_timeout", int(active_profile["read_timeout"])))
    if active_profile["write_timeout"]:
        assignments.append(("net_write_timeout", int(active_profile["write_timeout"])))
    if active_profile["max_execution_time"] != "":
        assignments.append(("max_execution_time", int(active_profile["max_execution_time"])))
    if active_profile["wait_timeout"]:
        assignments.append(("wait_timeout", int(active_profile["wait_timeout"])))
    if active_profile["interactive_timeout"]:
        assignments.append(("interactive_timeout", int(active_profile["interactive_timeout"])))
    if not assignments:
        return
    cursor = cnx.cursor()
    for variable_name, variable_value in assignments:
        cursor.execute(f"SET SESSION {variable_name} = {variable_value}")


class TunneledMySQLConnection:
    def __init__(self, connection, tunnel):
        object.__setattr__(self, "_connection", connection)
        object.__setattr__(self, "_tunnel", tunnel)

    def __getattr__(self, name):
        return getattr(self._connection, name)

    def __setattr__(self, name, value):
        if name in {"_connection", "_tunnel"}:
            object.__setattr__(self, name, value)
            return
        setattr(self._connection, name, value)

    def close(self):
        try:
            self._connection.close()
        finally:
            self._tunnel.stop()


class CachedMySQLConnection:
    def __init__(self, connection, cache_key):
        object.__setattr__(self, "_connection", connection)
        object.__setattr__(self, "_cache_key", cache_key)
        object.__setattr__(self, "_cached_mysql_connection", True)

    def __getattr__(self, name):
        return getattr(self._connection, name)

    def __setattr__(self, name, value):
        if name in {"_connection", "_cache_key", "_cached_mysql_connection"}:
            object.__setattr__(self, name, value)
            return
        setattr(self._connection, name, value)

    def close(self):
        return None

    def real_close(self):
        try:
            self._connection.close()
        except Exception:
            pass


def _close_connection_object(cnx):
    if not cnx:
        return
    try:
        if isinstance(cnx, CachedMySQLConnection):
            cnx.real_close()
        else:
            cnx.close()
    except Exception:
        pass


def close_mysql_connection(cnx):
    if not cnx:
        return
    if isinstance(cnx, CachedMySQLConnection):
        try:
            if not cnx.autocommit:
                cnx.rollback()
        except Exception:
            _purge_cached_mysql_connection(cnx._cache_key)
        return
    _close_connection_object(cnx)


def _connection_cache_session_id():
    if not session.get("logged_in"):
        return ""
    cache_id = session.get("connection_cache_id")
    if not cache_id:
        cache_id = secrets.token_urlsafe(18)
        session["connection_cache_id"] = cache_id
    return cache_id


def _connection_cache_signature(connection_config, profile):
    password_hash = hashlib.sha256(str(connection_config.get("password", "")).encode("utf-8")).hexdigest()
    payload = {
        "user": str(connection_config.get("user", "")),
        "password_hash": password_hash,
        "host": str(profile.get("host", "")),
        "port": int(profile.get("port") or 0),
        "database": str(connection_config.get("database", "")),
        "ssh_enabled": bool(profile.get("ssh_enabled")),
        "ssh_host": str(profile.get("ssh_host", "")),
        "ssh_port": int(profile.get("ssh_port") or 0),
        "ssh_user": str(profile.get("ssh_user", "")),
        "ssh_key_file": str(profile.get("ssh_key_file", "")),
        "read_timeout": str(connection_config.get("read_timeout", "")),
        "write_timeout": str(connection_config.get("write_timeout", "")),
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()


def _connection_cache_key(connection_config, profile):
    cache_id = _connection_cache_session_id()
    if not cache_id:
        return None
    return (
        cache_id,
        threading.get_ident(),
        _connection_cache_signature(connection_config, profile),
    )


def _cached_connection_is_alive(cnx):
    try:
        if time.time() - getattr(cnx, "_created_at", time.time()) > CONNECTION_CACHE_MAX_AGE_SECONDS:
            return False
        cnx.ping(reconnect=False, attempts=1, delay=0)
        return True
    except Exception:
        return False


def _purge_cached_mysql_connection(cache_key):
    with _CONNECTION_CACHE_LOCK:
        cached = _CONNECTION_CACHE.pop(cache_key, None)
    if cached:
        _close_connection_object(cached.get("connection"))


def clear_cached_mysql_connections_for_session():
    cache_id = session.get("connection_cache_id")
    if not cache_id:
        return
    with _CONNECTION_CACHE_LOCK:
        keys = [key for key in _CONNECTION_CACHE if key[0] == cache_id]
        cached_rows = [_CONNECTION_CACHE.pop(key) for key in keys]
    for row in cached_rows:
        _close_connection_object(row.get("connection"))


def _build_tunneled_connection_config(connection_config, profile):
    if not profile.get("ssh_enabled"):
        return connection_config, None
    key_file = os.path.expanduser(str(profile.get("ssh_key_file", "")).strip())
    if not key_file or not Path(key_file).exists():
        raise RuntimeError("SSH private key file does not exist: {}".format(profile.get("ssh_key_file", "")))
    try:
        from sshtunnel import SSHTunnelForwarder
    except Exception as error:
        raise RuntimeError("The sshtunnel package is not installed. Run setup.sh to refresh requirements.") from error

    tunnel = SSHTunnelForwarder(
        (profile["ssh_host"], int(profile["ssh_port"])),
        ssh_username=profile["ssh_user"],
        ssh_pkey=key_file,
        remote_bind_address=(connection_config["host"], int(connection_config["port"])),
        local_bind_address=("127.0.0.1", 0),
    )
    tunnel.start()
    tunneled_config = dict(connection_config)
    tunneled_config["host"] = "127.0.0.1"
    tunneled_config["port"] = int(tunnel.local_bind_port)
    return tunneled_config, tunnel


def mysql_connection(config=None, *, autocommit=False):
    profile = get_session_profile()
    connection_config = dict(config or get_connection_config())
    configured_connection_timeout = _normalized_optional_timeout(connection_config.get("connection_timeout"))
    connection_config["connection_timeout"] = configured_connection_timeout or DEFAULT_CONNECTION_TIMEOUT_SECONDS
    cache_key = _connection_cache_key(connection_config, profile)
    if cache_key:
        with _CONNECTION_CACHE_LOCK:
            cached_row = _CONNECTION_CACHE.get(cache_key)
        if cached_row:
            cached_connection = cached_row.get("connection")
            if _cached_connection_is_alive(cached_connection):
                cached_row["last_used_at"] = time.time()
                cached_connection.autocommit = bool(autocommit)
                _apply_connection_profile_session_settings(cached_connection)
                return cached_connection
            _purge_cached_mysql_connection(cache_key)

    cnx = None
    tunnel = None
    try:
        connection_config, tunnel = _build_tunneled_connection_config(connection_config, profile)
        cnx = mysql.connector.connect(**connection_config)
        if tunnel:
            cnx = TunneledMySQLConnection(cnx, tunnel)
        cnx.can_consume_results = True
        _apply_connection_profile_session_settings(cnx)
        cnx.autocommit = bool(autocommit)
        if cache_key:
            cnx = CachedMySQLConnection(cnx, cache_key)
            object.__setattr__(cnx, "_created_at", time.time())
            with _CONNECTION_CACHE_LOCK:
                _CONNECTION_CACHE[cache_key] = {
                    "connection": cnx,
                    "created_at": time.time(),
                    "last_used_at": time.time(),
                }
        return cnx
    except Exception:
        close_mysql_connection(cnx)
        if tunnel:
            try:
                tunnel.stop()
            except Exception:
                pass
        raise


def run_sql(sql_text, params=None, *, include_database=True, autocommit=False):
    cnx = None
    cursor = None
    try:
        cnx = mysql_connection(
            get_connection_config(include_database=include_database),
            autocommit=autocommit,
        )
        cursor = cnx.cursor()
        cursor.execute(sql_text, params or ())
        return cursor.fetchall()
    finally:
        close_mysql_connection(cnx)


def exec_sql(sql_text, params=None, *, include_database=True, autocommit=False):
    cnx = None
    cursor = None
    try:
        cnx = mysql_connection(
            get_connection_config(include_database=include_database),
            autocommit=autocommit,
        )
        cursor = cnx.cursor()
        cursor.execute(sql_text, params or ())
        cnx.commit()
        return True
    except mysql.connector.Error:
        raise
    finally:
        close_mysql_connection(cnx)


def run_sql_with_columns(sql_text, params=None, *, include_database=True, autocommit=False):
    cnx = None
    cursor = None
    try:
        cnx = mysql_connection(
            get_connection_config(include_database=include_database),
            autocommit=autocommit,
        )
        cursor = cnx.cursor()
        cursor.execute(sql_text, params or ())
        return {
            "columns": list(cursor.column_names or ()),
            "rows": [list(row) for row in cursor.fetchall()],
        }
    finally:
        close_mysql_connection(cnx)


def run_sql_multi_resultsets(sql_text, params=None, *, include_database=True, autocommit=False):
    cnx = None
    cursor = None
    try:
        if params:
            raise ValueError("Parameterized multi-result SQL is not supported.")
        cnx = mysql_connection(
            get_connection_config(include_database=include_database),
            autocommit=autocommit,
        )
        cursor = cnx.cursor()
        datasets = []
        cursor.execute(sql_text, map_results=True)
        result_index = 1
        while True:
            if cursor.with_rows:
                datasets.append(
                    {
                        "title": "Result Set {}".format(result_index),
                        "columns": list(cursor.column_names or ()),
                        "rows": [
                            [_normalize_modal_cell(value) for value in row]
                            for row in cursor.fetchall()
                        ],
                    }
                )
            else:
                datasets.append(
                    {
                        "title": "Status {}".format(result_index),
                        "columns": ["statement", "rowcount"],
                        "rows": [[
                            _normalize_modal_cell(getattr(cursor, "statement", "")),
                            _normalize_modal_cell(cursor.rowcount),
                        ]],
                    }
                )
            result_index += 1
            if not cursor.nextset():
                break
        cnx.commit()
        return datasets
    except (mysql.connector.Error, ValueError):
        raise
    finally:
        close_mysql_connection(cnx)


def run_sql_dicts(sql_text, params=None, *, include_database=True, autocommit=False):
    result = run_sql_with_columns(
        sql_text,
        params=params,
        include_database=include_database,
        autocommit=autocommit,
    )
    return [dict(zip(result["columns"], row)) for row in result["rows"]]


def call_proc(proc_name, args):
    cnx = None
    cursor = None
    try:
        cnx = mysql_connection()
        cursor = cnx.cursor()
        result_args = cursor.callproc(proc_name, args)
        datasets = []
        columns = []
        for result in cursor.stored_results():
            datasets.append(result.fetchall())
            columns.append(result.column_names)
        return {
            "output": result_args[1] if len(result_args) > 1 else "",
            "resultset": datasets,
            "columnset": columns,
        }
    finally:
        close_mysql_connection(cnx)


def _normalize_modal_cell(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def login_required(route_handler):
    @wraps(route_handler)
    def wrapped(*args, **kwargs):
        if not session.get("logged_in"):
            flash("Log in with a saved connection profile first.", "warning")
            return redirect(url_for("login"))
        return route_handler(*args, **kwargs)

    return wrapped


@app.route("/connection-timeouts", methods=["POST"])
@login_required
def update_connection_timeouts():
    current_profile = get_session_profile()
    if not current_profile.get("name"):
        flash("Choose a saved connection profile before changing timeout settings.", "warning")
        return redirect(request.referrer or url_for("home"))

    updated_profile = dict(current_profile)
    updated_profile["connection_timeout"] = _normalized_optional_timeout(request.form.get("connection_timeout"))
    updated_profile["read_timeout"] = _normalized_optional_timeout(request.form.get("read_timeout"))
    updated_profile["write_timeout"] = _normalized_optional_timeout(request.form.get("write_timeout"))
    updated_profile["max_execution_time"] = _normalized_optional_timeout(
        request.form.get("max_execution_time"),
        allow_zero=True,
    )
    updated_profile["wait_timeout"] = _normalized_optional_timeout(request.form.get("wait_timeout"))
    updated_profile["interactive_timeout"] = _normalized_optional_timeout(request.form.get("interactive_timeout"))

    existing_profiles = load_profiles()
    save_profiles(
        [
            updated_profile if row["name"].lower() == updated_profile["name"].lower() else row
            for row in existing_profiles
        ]
    )
    set_session_profile(updated_profile)
    flash("Connection timeout settings updated.", "success")
    return redirect(request.referrer or url_for("home"))


def validate_user(user, password):
    config = get_connection_config(
        user=user,
        password=password,
        include_database=bool(get_session_profile()["database"]),
        fallback_connection_timeout=LOGIN_VALIDATION_CONNECTION_TIMEOUT_SECONDS,
    )
    try:
        cnx = mysql_connection(config)
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            return False, "Something is wrong with your user name or password."
        if err.errno == errorcode.ER_BAD_DB_ERROR:
            return False, "The configured default database does not exist."
        return False, str(err)
    cnx.close()
    return True, ""


def _is_public_endpoint(endpoint):
    return endpoint in {
        "login",
        "save_profile_route",
        "logout",
        "static",
    }


def _validate_active_session_connection():
    cnx = None
    cursor = None
    try:
        cnx = mysql_connection(
            get_connection_config(
                include_database=bool(get_session_profile()["database"]),
                fallback_connection_timeout=SESSION_VALIDATION_CONNECTION_TIMEOUT_SECONDS,
                fallback_read_timeout=SESSION_VALIDATION_CONNECTION_TIMEOUT_SECONDS,
                fallback_write_timeout=SESSION_VALIDATION_CONNECTION_TIMEOUT_SECONDS,
            )
        )
        cursor = cnx.cursor()
        cursor.execute("select 1")
        cursor.fetchone()
        return True
    except mysql.connector.Error:
        return False
    finally:
        close_mysql_connection(cnx)


@app.before_request
def enforce_live_database_session():
    if not session.get("logged_in"):
        return None
    if _is_public_endpoint(request.endpoint):
        return None
    if _validate_active_session_connection():
        return None
    clear_login_state()
    flash(
        "The MySQL connection is no longer available. Log in again after the database is back online.",
        "warning",
    )
    return redirect(url_for("login"))


def setup_db():
    exec_sql("create database if not exists nlsql", include_database=False)
    exec_sql(
        """
        create table if not exists nlsql.configdb (
            db_name varchar(64) not null primary key,
            enabled char(1) not null
        )
        """,
        include_database=False,
    )
    exec_sql(
        """
        insert into nlsql.configdb (db_name, enabled)
        select defaults.schema_name_value, 'Y'
        from (
            select 'information_schema' as schema_name_value
            union all
            select 'sys'
            union all
            select 'performance_schema'
        ) defaults
        left join nlsql.configdb existing
            on existing.db_name = defaults.schema_name_value
        where existing.db_name is null
        """,
        include_database=False,
    )


def fetch_enabled_databases():
    rows = run_sql(
        """
        select db_name as db_name_value
        from nlsql.configdb
        where enabled = 'Y'
        order by db_name
        """,
        include_database=False,
    )
    return [row[0] for row in rows]


def fetch_available_databases():
    rows = run_sql(
        """
        select schema_name as schema_name_value
        from information_schema.schemata
        where schema_name <> 'nlsql'
        order by schema_name
        """,
        include_database=False,
    )
    return [row[0] for row in rows]


def save_configdb_databases(database_names):
    cnx = None
    cursor = None
    try:
        cnx = mysql_connection(get_connection_config(include_database=False))
        cursor = cnx.cursor()
        cursor.execute("delete from nlsql.configdb")
        if database_names:
            cursor.executemany(
                """
                insert into nlsql.configdb (db_name, enabled)
                values (%s, 'Y')
                """,
                [(name,) for name in sorted(set(database_names))],
            )
        cnx.commit()
    except mysql.connector.Error:
        raise
    finally:
        close_mysql_connection(cnx)


def setup_askme_db():
    exec_sql(
        "create database if not exists {}".format(_quote_identifier(ASKME_SCHEMA_NAME)),
        include_database=False,
    )
    exec_sql(
        """
        create table if not exists askme.config (
            my_row_id bigint unsigned not null auto_increment,
            env_var varchar(128) not null,
            env_value text null,
            primary key (my_row_id)
        )
        """,
        include_database=False,
    )
    existing_columns = run_sql(
        """
        select column_name as column_name_value
        from information_schema.columns
        where table_schema = 'askme'
          and table_name = 'config'
        order by ordinal_position
        """,
        include_database=False,
    )
    column_names = [str(row[0]).lower() for row in existing_columns]
    if "my_row_id" not in column_names:
        exec_sql(
            """
            alter table askme.config
                add column my_row_id bigint unsigned not null auto_increment first,
                drop primary key,
                add primary key (my_row_id)
            """,
            include_database=False,
        )
    existing_indexes = run_sql(
        """
        select index_name as index_name_value
        from information_schema.statistics
        where table_schema = 'askme'
          and table_name = 'config'
        """,
        include_database=False,
    )
    index_names = {str(row[0]).lower() for row in existing_indexes}
    if "uk_askme_config_env_var" in index_names:
        exec_sql(
            "alter table askme.config drop index uk_askme_config_env_var",
            include_database=False,
        )
    for env_var, default_value in ASKME_CONFIG_ROWS:
        exec_sql(
            """
            insert into askme.config (env_var, env_value)
            select %s, %s
            from dual
            where not exists (
                select 1
                from askme.config
                where env_var = %s
            )
            """,
            (env_var, default_value, env_var),
            include_database=False,
        )


def fetch_askme_config():
    rows = run_sql(
        """
        select env_var as env_var_value, coalesce(env_value, '') as env_value_value
        from askme.config
        order by env_var
        """,
        include_database=False,
    )
    return {str(row[0]): str(row[1] or "") for row in rows}


def askme_setup_is_ready():
    try:
        config_values = fetch_askme_config()
    except mysql.connector.Error:
        return False
    required_keys = {env_var for env_var, _default_value in ASKME_CONFIG_ROWS}
    for key in required_keys:
        if not str(config_values.get(key, "")).strip():
            return False
    return True


def save_askme_config(config_values):
    cnx = None
    cursor = None
    try:
        cnx = mysql_connection(get_connection_config(include_database=False))
        cursor = cnx.cursor()
        cursor.execute("delete from askme.config")
        cursor.executemany(
            """
            insert into askme.config (env_var, env_value)
            values (%s, %s)
            """,
            [
                (env_var, str(config_values.get(env_var, "")).strip())
                for env_var, _default_value in ASKME_CONFIG_ROWS
            ],
        )
        cnx.commit()
    except mysql.connector.Error:
        raise
    finally:
        close_mysql_connection(cnx)


def import_oci_sdk():
    try:
        return importlib.import_module("oci")
    except Exception as error:
        raise RuntimeError("The OCI SDK is not available. Install the project requirements first.") from error


def _config_lookup(config_values, *keys):
    for key in keys:
        value = str(config_values.get(key, "") or "").strip()
        if value:
            return value
    return ""


def get_oci_config_settings(config_values):
    config_file = _config_lookup(config_values, "OCI_CONFIG_FILE", "config_file") or "~/.oci/config"
    config_profile = _config_lookup(config_values, "OCI_CONFIG_PROFILE", "config_profile") or "DEFAULT"
    return {
        "config_file": os.path.expanduser(config_file),
        "config_profile": config_profile,
    }


def get_oci_object_storage_client(config_values, *, timeout=(30, 600)):
    oci = import_oci_sdk()
    settings = get_oci_config_settings(config_values)
    config_file = settings["config_file"]
    config_profile = settings["config_profile"]
    region_value = _config_lookup(config_values, "OCI_REGION", "region")

    if config_file and Path(config_file).exists():
        try:
            oci_config = oci.config.from_file(file_location=config_file, profile_name=config_profile)
            if region_value:
                oci_config["region"] = region_value
            oci.config.validate_config(oci_config)
            return oci.object_storage.ObjectStorageClient(
                oci_config,
                retry_strategy=oci.retry.NoneRetryStrategy(),
                timeout=timeout,
            )
        except Exception as error:
            raise RuntimeError(
                "OCI config authentication failed for {} profile {}.".format(config_file, config_profile)
            ) from error

    try:
        signer = oci.auth.signers.InstancePrincipalsSecurityTokenSigner()
        return oci.object_storage.ObjectStorageClient(
            config={"region": region_value},
            signer=signer,
            retry_strategy=oci.retry.NoneRetryStrategy(),
            timeout=timeout,
        )
    except Exception as error:
        raise RuntimeError(
            "OCI authentication failed. Configure a readable ~/.oci/config profile or run on OCI Compute with instance-principal access."
        ) from error


def _get_model_ids(query):
    rows = run_sql(query)
    return [row[0] for row in rows]


def _get_cached_supported_llms(capability):
    normalized_capability = str(capability or "").strip().upper()
    if normalized_capability not in {"GENERATION", "TEXT_EMBEDDINGS"}:
        return []
    cache = session.get("llm_model_cache", {})
    if not isinstance(cache, dict):
        cache = {}
    cached_models = cache.get(normalized_capability)
    if isinstance(cached_models, list):
        return [str(model_id) for model_id in cached_models if str(model_id).strip()]

    rows = run_sql(
        """
        select model_id as model_id_value
        from sys.ML_SUPPORTED_LLMS
        where capabilities->>'$[0]' = %s
        order by model_id
        """,
        (normalized_capability,),
        include_database=False,
    )
    models = [str(row[0]).strip() for row in rows if row and row[0]]
    cache[normalized_capability] = models
    session["llm_model_cache"] = cache
    return models


def get_generation_models():
    return _get_cached_supported_llms("GENERATION")


def get_embedding_models():
    return _get_cached_supported_llms("TEXT_EMBEDDINGS")


def choose_default_model(models):
    if PREFERRED_DEFAULT_MODEL in models:
        return PREFERRED_DEFAULT_MODEL
    return models[0] if models else ""


def _quote_identifier(identifier):
    return "`{}`".format(str(identifier).replace("`", "``"))


def _pick_present_column(columns, candidates):
    column_lookup = {column.lower(): column for column in columns}
    for candidate in candidates:
        resolved = column_lookup.get(candidate.lower())
        if resolved:
            return resolved
    return ""


def _pick_memory_columns(columns):
    preferred = [
        "memory_used",
        "memory_usage",
        "used_memory",
        "total_memory",
        "free_memory",
        "memory_size",
        "ram_size",
        "ram_usage",
    ]
    matches = []
    for candidate in preferred:
        resolved = _pick_present_column(columns, [candidate])
        if resolved and resolved not in matches:
            matches.append(resolved)
    for column in columns:
        if "memory" in column.lower() and column not in matches:
            matches.append(column)
    return matches


def _unique(values):
    result = []
    seen = set()
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _normalize_progress(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip().replace(",", "")
        if text.endswith("%"):
            text = text[:-1].strip()
        try:
            number = float(text)
        except ValueError:
            return None
    if 0 <= number <= 1:
        return number * 100
    return number


def _format_bytes(value):
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        return str(value or "0 B")
    units = ["B", "KB", "MB", "GB", "TB", "PB"]
    unit_index = 0
    while number >= 1024 and unit_index < len(units) - 1:
        number /= 1024
        unit_index += 1
    if unit_index == 0:
        return f"{int(number)} {units[unit_index]}"
    return f"{number:.2f} {units[unit_index]}"


def _format_progress(value):
    if value is None:
        return "Unavailable"
    if abs(value - round(value)) < 0.001:
        return f"{int(round(value))}%"
    return f"{value:.1f}%"


def _split_heatwave_object_name(raw_name):
    normalized = str(raw_name or "").strip()
    if not normalized:
        return "", ""
    cleaned = normalized.replace("`", "")
    if "." in cleaned:
        left, right = cleaned.split(".", 1)
        return left.strip(), right.strip()
    return "", cleaned.strip()


def _derive_heatwave_row_class(progress_value, status_text="", error_text=""):
    status_value = str(status_text or "").strip().lower()
    error_value = str(error_text or "").strip()
    if error_value or any(token in status_value for token in ("error", "fail", "abort")):
        return "error"
    if progress_value is not None and progress_value >= 100:
        return "loaded"
    return "partial"


def _sanitize_import_column_name(value, index, seen_names):
    text = str(value or "").strip()
    if not text:
        text = f"column_{index}"
    normalized = re.sub(r"[^A-Za-z0-9_$]+", "_", text).strip("_")
    if not normalized:
        normalized = f"column_{index}"
    if not re.match(r"^[A-Za-z_]", normalized):
        normalized = f"column_{index}_{normalized}"
    normalized = normalized[:64]
    candidate = normalized
    suffix = 1
    while candidate.lower() in seen_names:
        suffix_text = f"_{suffix}"
        candidate = f"{normalized[: max(1, 64 - len(suffix_text))]}{suffix_text}"
        suffix += 1
    seen_names.add(candidate.lower())
    return _validate_identifier(candidate, "Column name")


def _default_import_table_name_from_filename(filename):
    stem = Path(str(filename or "").strip()).stem
    return _sanitize_import_column_name(stem, 1, set())


def _normalize_import_cell(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    text = str(value)
    return text if text != "" else None


def _read_csv_import(file_storage):
    file_storage.stream.seek(0)
    text_stream = io.TextIOWrapper(file_storage.stream, encoding="utf-8-sig", newline="")
    try:
        rows = [list(row) for row in csv.reader(text_stream)]
    finally:
        try:
            text_stream.detach()
        except Exception:
            pass
    return rows


def _read_csv_import_path(file_path):
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
        return [list(row) for row in csv.reader(handle)]


def _read_excel_import(file_storage):
    file_storage.stream.seek(0)
    workbook = load_workbook(file_storage.stream, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_excel_import_path(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _normalize_import_dataset(rows, filename):
    non_empty_rows = [
        [None if value is None else value for value in row]
        for row in rows
        if any(str(value).strip() for value in row if value is not None)
    ]
    if not non_empty_rows:
        raise ValueError("The selected file does not contain any rows.")

    raw_headers = non_empty_rows[0]
    data_rows = non_empty_rows[1:]
    if not any(str(value).strip() for value in raw_headers if value is not None):
        raise ValueError("The first row must contain column names.")

    seen_names = set()
    headers = [
        _sanitize_import_column_name(value, index + 1, seen_names)
        for index, value in enumerate(raw_headers)
    ]
    normalized_rows = []
    for row in data_rows:
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        normalized_rows.append([_normalize_import_cell(value) for value in padded[: len(headers)]])

    return {
        "filename": filename,
        "headers": headers,
        "rows": normalized_rows,
        "row_count": len(normalized_rows),
    }


def _load_import_rows(file_storage):
    filename = str(file_storage.filename or "").strip()
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        rows = _read_csv_import(file_storage)
    else:
        raise ValueError("Choose a CSV file to import.")
    return _normalize_import_dataset(rows, filename)


def _ensure_import_preview_dir():
    IMPORT_PREVIEW_DIR.mkdir(parents=True, exist_ok=True)


def _save_import_preview_file(file_storage):
    filename = str(file_storage.filename or "").strip()
    extension = Path(filename).suffix.lower()
    if extension != ".csv":
        raise ValueError("Choose a CSV file to import.")
    _ensure_import_preview_dir()
    token = secrets.token_hex(16)
    target_path = IMPORT_PREVIEW_DIR / f"{token}{extension}"
    file_storage.stream.seek(0)
    with open(target_path, "wb") as handle:
        handle.write(file_storage.stream.read())
    return token, target_path


def _resolve_import_preview_path(token):
    normalized = str(token or "").strip()
    if not re.fullmatch(r"[0-9a-f]{32}", normalized):
        raise ValueError("The loaded import preview is invalid or expired.")
    _ensure_import_preview_dir()
    matches = list(IMPORT_PREVIEW_DIR.glob(f"{normalized}.*"))
    if not matches:
        raise ValueError("The loaded import preview is invalid or expired.")
    return matches[0]


def _load_import_rows_from_path(token):
    file_path = _resolve_import_preview_path(token)
    extension = file_path.suffix.lower()
    if extension == ".csv":
        rows = _read_csv_import_path(file_path)
    else:
        raise ValueError("The loaded import preview is invalid or expired.")
    return _normalize_import_dataset(rows, file_path.name)


def _delete_import_preview_file(token):
    try:
        file_path = _resolve_import_preview_path(token)
    except ValueError:
        return
    try:
        file_path.unlink()
    except OSError:
        pass


def _build_import_preview_table(import_payload, max_rows=50):
    headers = list(import_payload.get("headers") or [])
    rows = list(import_payload.get("rows") or [])
    return {
        "columns": headers,
        "column_definitions": _build_import_table_columns(headers, rows) if headers else [],
        "rows": [list(row) for row in rows[:max_rows]],
        "row_count": int(import_payload.get("row_count") or 0),
        "preview_count": min(int(import_payload.get("row_count") or 0), max_rows),
        "filename": str(import_payload.get("filename") or ""),
    }


def _infer_import_column_type(values):
    non_empty = [str(value).strip() for value in values if value not in (None, "")]
    if not non_empty:
        return "VARCHAR(255)"
    max_length = max(len(value) for value in non_empty)
    if max_length <= 255:
        return f"VARCHAR({max(1, min(255, max_length))})"
    if max_length <= 65535:
        return "TEXT"
    return "LONGTEXT"


def _build_import_table_columns(headers, rows):
    columns = []
    for index, header in enumerate(headers):
        sample_values = [row[index] for row in rows]
        columns.append(
            {
                "name": header,
                "data_type": _infer_import_column_type(sample_values),
                "nullable": True,
                "primary": False,
            }
        )
    return columns


def _normalize_import_primary_key_mode(value, *, add_invisible_primary_key=False):
    normalized = str(value or "").strip().lower()
    if not normalized:
        normalized = "my_row_id" if add_invisible_primary_key else "none"
    if normalized not in {"none", "columns", "my_row_id"}:
        raise ValueError("Choose a valid primary key option.")
    return normalized


def _apply_import_primary_key_definition(columns, headers, rows, primary_key_mode, primary_key_columns):
    normalized_mode = _normalize_import_primary_key_mode(primary_key_mode)
    if normalized_mode == "none":
        return False
    if normalized_mode == "my_row_id":
        if any(str(column.get("name", "")).strip().lower() == "my_row_id" for column in columns):
            raise ValueError("Column name `my_row_id` is reserved for the generated invisible primary key.")
        return True

    selected_columns = [str(column or "").strip() for column in primary_key_columns or [] if str(column or "").strip()]
    if not selected_columns:
        raise ValueError("Select at least one import column for the primary key, or choose `my_row_id`.")

    header_lookup = {header.lower(): header for header in headers}
    normalized_selected_columns = []
    seen_selected = set()
    for selected_column in selected_columns:
        normalized_selected = header_lookup.get(selected_column.lower())
        if not normalized_selected:
            raise ValueError(f"Primary key column `{selected_column}` was not found in the import file.")
        if normalized_selected.lower() in seen_selected:
            continue
        seen_selected.add(normalized_selected.lower())
        normalized_selected_columns.append(normalized_selected)

    selected_indexes = [headers.index(column_name) for column_name in normalized_selected_columns]
    seen_values = set()
    for row_number, row in enumerate(rows, start=2):
        key_values = tuple(row[index] for index in selected_indexes)
        if any(value in (None, "") for value in key_values):
            raise ValueError(f"Primary key columns cannot contain empty values. Check CSV row {row_number}.")
        if key_values in seen_values:
            raise ValueError(f"Primary key columns must be unique. Duplicate key found at CSV row {row_number}.")
        seen_values.add(key_values)

    selected_lookup = {column_name.lower() for column_name in normalized_selected_columns}
    for column in columns:
        if str(column.get("name", "")).lower() in selected_lookup:
            column["primary"] = True
            column["nullable"] = False
    return False


def _normalize_import_row_limit(value):
    normalized = str(value or "full").strip().lower()
    if normalized not in {"full", "1000"}:
        raise ValueError("Choose a valid import row option.")
    return normalized


def import_file_to_table(
    schema_name,
    table_name,
    import_payload,
    *,
    overwrite_existing=False,
    create_new_table=False,
    create_schema=False,
    add_invisible_primary_key=False,
    primary_key_mode=None,
    primary_key_columns=None,
    import_row_limit="full",
):
    database_name = _validate_identifier(schema_name, "Database name")
    normalized_table_name = _validate_identifier(table_name, "Table name")
    row_limit = _normalize_import_row_limit(import_row_limit)
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be modified.")
    database_exists = _database_exists(database_name)
    if not database_exists and create_schema:
        create_database(database_name)
        database_exists = True
    if not database_exists:
        raise ValueError(f"Database `{database_name}` does not exist.")
    headers = list(import_payload.get("headers") or [])
    all_rows = list(import_payload.get("rows") or [])
    rows = all_rows[:1000] if row_limit == "1000" else all_rows
    if not headers:
        raise ValueError("The import file must include a header row.")

    table_exists = _table_exists(database_name, normalized_table_name)
    if table_exists and create_new_table:
        raise ValueError(f"Table `{database_name}.{normalized_table_name}` already exists. Choose another table name.")

    if not table_exists:
        import_columns = _build_import_table_columns(headers, rows)
        if primary_key_mode is None:
            primary_key_mode = "my_row_id" if add_invisible_primary_key else "none"
        add_generated_primary_key = _apply_import_primary_key_definition(
            import_columns,
            headers,
            rows,
            primary_key_mode,
            primary_key_columns,
        )
        create_table(
            database_name,
            normalized_table_name,
            import_columns,
            add_invisible_auto_pk=add_generated_primary_key,
        )
    else:
        existing_columns = {row["column_name"].lower(): row["column_name"] for row in fetch_table_definition(database_name, normalized_table_name)}
        missing_columns = [header for header in headers if header.lower() not in existing_columns]
        if missing_columns:
            raise ValueError(
                "Import file columns were not found in the existing table: {}.".format(", ".join(missing_columns))
            )
        headers = [existing_columns[header.lower()] for header in headers]
        if overwrite_existing:
            exec_sql(
                "truncate table {}.{}".format(
                    _quote_identifier(database_name),
                    _quote_identifier(normalized_table_name),
                ),
                include_database=False,
            )

    if rows:
        placeholders = ", ".join(["%s"] * len(headers))
        cnx = None
        cursor = None
        try:
            cnx = mysql_connection(get_connection_config(include_database=False))
            cursor = cnx.cursor()
            cursor.executemany(
                "insert into {}.{} ({}) values ({})".format(
                    _quote_identifier(database_name),
                    _quote_identifier(normalized_table_name),
                    ", ".join(_quote_identifier(header) for header in headers),
                    placeholders,
                ),
                rows,
            )
            cnx.commit()
        except mysql.connector.Error:
            raise
        finally:
            close_mysql_connection(cnx)

    return {
        "database_name": database_name,
        "table_name": normalized_table_name,
        "created_table": not table_exists,
        "overwrite_existing": bool(table_exists and overwrite_existing),
        "row_count": len(rows),
        "source_row_count": len(all_rows),
        "import_row_limit": row_limit,
        "column_count": len(headers),
        "filename": import_payload.get("filename", ""),
    }


def _format_uptime(seconds):
    try:
        total_seconds = int(float(seconds or 0))
    except (TypeError, ValueError):
        return str(seconds or "-")
    days, remainder = divmod(total_seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, secs = divmod(remainder, 60)
    parts = []
    if days:
        parts.append(f"{days}d")
    if days or hours:
        parts.append(f"{hours}h")
    if days or hours or minutes:
        parts.append(f"{minutes}m")
    parts.append(f"{secs}s")
    return " ".join(parts)


def _validate_identifier(identifier, label):
    value = str(identifier or "").strip()
    if not value:
        raise ValueError(f"{label} is required.")
    if not IDENTIFIER_RE.match(value):
        raise ValueError(f"{label} must use only letters, digits, `_`, or `$`.")
    return value


def _normalize_mysql_data_type(data_type):
    value = str(data_type or "").strip().upper()
    value = re.sub(r"\s+", " ", value)
    if not value:
        raise ValueError("Column data type is required.")
    if not MYSQL_DATA_TYPE_RE.match(value):
        raise ValueError(f"Unsupported MySQL data type: {data_type}.")
    return value


def _build_mysql_data_type(base_type, params_text=""):
    normalized_base_type = str(base_type or "").strip().upper()
    option = MYSQL_TYPE_OPTION_MAP.get(normalized_base_type)
    if not option:
        raise ValueError(f"Unsupported MySQL data type: {base_type}.")

    cleaned_params = str(params_text or "").strip()
    if option["param_mode"] == "none":
        return normalized_base_type

    if not cleaned_params:
        raise ValueError(f"{normalized_base_type} requires size parameters.")

    if option["param_mode"] == "single":
        if not re.fullmatch(r"\d+", cleaned_params):
            raise ValueError(f"{normalized_base_type} size must be a single integer.")
    elif option["param_mode"] == "multi":
        if not re.fullmatch(r"\d+\s*,\s*\d+", cleaned_params):
            raise ValueError(f"{normalized_base_type} size must use two integers, for example 10,3.")

    return _normalize_mysql_data_type(f"{normalized_base_type}({cleaned_params})")


def _parse_secondary_engine(create_options):
    text = str(create_options or "").upper()
    match = re.search(r'SECONDARY_ENGINE="?([A-Z0-9_]+)"?', text)
    return match.group(1) if match else ""


def _split_mysql_data_type(column_type, data_type):
    base_type = str(data_type or "").strip().upper()
    raw_column_type = str(column_type or "").strip()
    params = ""
    match = re.search(r"\(([^()]*)\)", raw_column_type)
    if match:
        params = match.group(1).strip()
    if not base_type:
        normalized_column_type = _normalize_mysql_data_type(raw_column_type)
        base_match = re.match(r"^([A-Z]+(?:\s+[A-Z]+)?)", normalized_column_type)
        base_type = base_match.group(1) if base_match else normalized_column_type
        params_match = re.search(r"\(([^()]*)\)", normalized_column_type)
        params = params_match.group(1).strip() if params_match else params
    return base_type, params


def _table_exists(schema_name, table_name):
    rows = run_sql(
        """
        select count(*) as row_count
        from information_schema.tables
        where table_schema = %s
          and table_name = %s
        """,
        (schema_name, table_name),
        include_database=False,
    )
    return bool(rows and rows[0][0])


def _get_table_columns(schema_name, table_name):
    rows = run_sql(
        """
        select column_name as column_name_value
        from information_schema.columns
        where table_schema = %s
          and table_name = %s
        order by ordinal_position
        """,
        (schema_name, table_name),
        include_database=False,
    )
    return [row[0] for row in rows]


def _database_exists(schema_name):
    rows = run_sql(
        """
        select schema_name as schema_name_value
        from information_schema.schemata
        where schema_name = %s
        """,
        (schema_name,),
        include_database=False,
    )
    return bool(rows)


def _is_system_database(schema_name):
    name = str(schema_name or "")
    return name in SYSTEM_DATABASES or name.startswith("mysql_")


def _default_table_form():
    return {
        "table_name": "",
        "columns": [
            {
                "name": "",
                "type_name": "VARCHAR",
                "type_params": "255",
                "nullable": True,
                "primary": False,
            }
        ],
    }


def _default_column_form():
    return {
        "original_name": "",
        "name": "",
        "type_name": "VARCHAR",
        "type_params": "255",
        "nullable": True,
    }


def _default_import_form():
    return {
        "database_name": "",
        "new_schema": False,
        "new_database_name": "",
        "table_name": "",
        "overwrite_existing": False,
        "create_new_table": False,
        "add_invisible_primary_key": False,
        "primary_key_mode": "my_row_id",
        "primary_key_columns": [],
        "import_row_limit": "full",
        "preview_token": "",
    }


def _normalize_db_admin_tab(value):
    tab = str(value or "db").strip().lower()
    if tab in DB_ADMIN_MONITORING_VIEWS:
        return "monitoring"
    return tab if tab in {item["id"] for item in DB_ADMIN_TABS} else "db"


def fetch_database_inventory():
    rows = run_sql(
        """
        select
            s.schema_name as database_name_value,
            coalesce(stats.table_count, 0) as table_count_value
        from information_schema.schemata s
        left join (
            select
                table_schema as table_schema_value,
                count(*) as table_count
            from information_schema.tables
            where table_type = 'BASE TABLE'
            group by table_schema
        ) stats
            on stats.table_schema_value = s.schema_name
        order by s.schema_name
        """,
        include_database=False,
    )
    inventory = [
        {
            "database_name": row[0],
            "table_count": int(row[1] or 0),
            "is_system": _is_system_database(row[0]),
        }
        for row in rows
    ]
    return sorted(
        inventory,
        key=lambda row: (0 if row["is_system"] else 1, row["database_name"].lower()),
    )


def fetch_import_tree():
    rows = run_sql(
        """
        select
            s.schema_name as database_name_value,
            t.table_name as table_name_value
        from information_schema.schemata s
        left join information_schema.tables t
            on t.table_schema = s.schema_name
           and t.table_type = 'BASE TABLE'
        order by s.schema_name, t.table_name
        """,
        include_database=False,
    )
    databases = {}
    for database_name, table_name in rows:
        entry = databases.setdefault(
            database_name,
            {
                "database_name": database_name,
                "is_system": _is_system_database(database_name),
                "tables": [],
            },
        )
        if table_name:
            entry["tables"].append(table_name)
    result = list(databases.values())
    result.sort(key=lambda row: (0 if row["is_system"] else 1, row["database_name"].lower()))
    return result


def fetch_tables_for_database(schema_name):
    rows = run_sql(
        """
        select
            t.table_name as table_name_value,
            coalesce(t.table_rows, 0) as row_count_value,
            coalesce(c.column_count, 0) as column_count_value,
            coalesce(t.engine, '') as engine_value,
            coalesce(t.create_options, '') as create_options_value
        from information_schema.tables t
        left join (
            select
                table_schema as table_schema_value,
                table_name as table_name_value,
                count(*) as column_count
            from information_schema.columns
            where table_schema = %s
            group by table_schema, table_name
        ) c
            on c.table_schema_value = t.table_schema
           and c.table_name_value = t.table_name
        where t.table_schema = %s
          and t.table_type = 'BASE TABLE'
        order by t.table_name
        """,
        (schema_name, schema_name),
        include_database=False,
    )
    heatwave_status = fetch_heatwave_load_status_for_database(schema_name)
    return [
        {
            "table_name": row[0],
            "row_count": int(row[1] or 0),
            "column_count": int(row[2] or 0),
            "engine": row[3] or "-",
            "secondary_engine": _parse_secondary_engine(row[4]),
            "engine_display": "{} / {}".format(
                row[3] or "-",
                _parse_secondary_engine(row[4]) or "-",
            ),
            "heatwave": heatwave_status.get(
                str(row[0]).lower(),
                {
                    "progress_value": None,
                    "progress_display": "-",
                    "is_loaded": False,
                },
            ),
        }
        for row in rows
    ]


def _extract_table_name_from_heatwave_row(schema_name, row):
    schema_value = str(row.get("schema_name", "") or "").strip()
    table_value = str(row.get("table_name", "") or "").strip()
    if schema_value and table_value and schema_value.lower() == schema_name.lower():
        return table_value

    composite_name = str(row.get("full_name", "") or "").strip()
    if composite_name:
        normalized = composite_name.strip("`")
        if "." in normalized:
            left, right = normalized.split(".", 1)
            if left.strip("`").lower() == schema_name.lower():
                return right.strip("`")
        elif normalized:
            return normalized

    if table_value:
        return table_value
    return ""


def fetch_heatwave_load_status_for_database(schema_name):
    if not (_table_exists("performance_schema", "rpd_tables") and _table_exists("performance_schema", "rpd_table_id")):
        return {}

    table_columns = _get_table_columns("performance_schema", "rpd_tables")
    table_id_columns = _get_table_columns("performance_schema", "rpd_table_id")

    rpd_table_id_column = _pick_present_column(table_columns, ["rpd_table_id", "table_id", "id"])
    progress_column = _pick_present_column(
        table_columns,
        ["load_progress", "loading_progress", "load_percentage", "pct_loaded", "progress"],
    )
    id_column = _pick_present_column(table_id_columns, ["id"])
    name_column = _pick_present_column(table_id_columns, ["name"])
    schema_column = _pick_present_column(table_id_columns, ["schema_name", "table_schema", "database_name", "db_name"])
    table_name_column = _pick_present_column(table_id_columns, ["table_name", "name"])

    if not (rpd_table_id_column and id_column and progress_column):
        return {}

    select_parts = [
        "t.{} as load_progress".format(_quote_identifier(progress_column)),
    ]
    if schema_column:
        select_parts.append("i.{} as schema_name".format(_quote_identifier(schema_column)))
    if table_name_column:
        select_parts.append("i.{} as table_name".format(_quote_identifier(table_name_column)))
    if name_column and name_column != table_name_column:
        select_parts.append("i.{} as full_name".format(_quote_identifier(name_column)))
    elif name_column:
        select_parts.append("i.{} as full_name".format(_quote_identifier(name_column)))

    query = """
    select {columns}
    from performance_schema.rpd_tables t
    inner join performance_schema.rpd_table_id i
        on i.{id_column} = t.{table_id_column}
    """.format(
        columns=", ".join(select_parts),
        id_column=_quote_identifier(id_column),
        table_id_column=_quote_identifier(rpd_table_id_column),
    )
    rows = run_sql_dicts(query, include_database=False)

    status_map = {}
    for row in rows:
        table_name = _extract_table_name_from_heatwave_row(schema_name, row)
        if not table_name:
            continue
        progress_value = _normalize_progress(row.get("load_progress"))
        if progress_value is None:
            continue
        existing = status_map.get(table_name.lower())
        if existing is None or progress_value > existing["progress_value"]:
            status_map[table_name.lower()] = {
                "progress_value": progress_value,
                "progress_display": _format_progress(progress_value),
                "is_loaded": progress_value >= 100,
            }
    return status_map


def fetch_heatwave_tables_report():
    report = {
        "columns": [
            "Schema",
            "Table",
            "HeatWave Status",
            "Load Progress",
            "Status Detail",
            "Error Detail",
            "Size Bytes",
            "Query Count",
            "Recovery Source",
            "Load Start",
            "Duration (sec)",
        ],
        "rows": [],
        "row_classes": [],
        "row_actions": [],
    }
    if not (_table_exists("performance_schema", "rpd_tables") and _table_exists("performance_schema", "rpd_table_id")):
        return report

    table_columns = _get_table_columns("performance_schema", "rpd_tables")
    table_id_columns = _get_table_columns("performance_schema", "rpd_table_id")

    table_id_column = _pick_present_column(table_columns, ["rpd_table_id", "table_id", "id"])
    id_column = _pick_present_column(table_id_columns, ["id"])
    name_column = _pick_present_column(table_id_columns, ["table_name", "name"])
    schema_column = _pick_present_column(table_id_columns, ["schema_name", "table_schema", "database_name", "db_name"])
    full_name_column = _pick_present_column(table_id_columns, ["name"])
    progress_column = _pick_present_column(
        table_columns,
        ["load_progress", "loading_progress", "load_percentage", "pct_loaded", "progress"],
    )
    status_column = _pick_present_column(table_columns, ["status", "load_status", "state"])
    error_column = _pick_present_column(table_columns, ["error_message", "error", "last_error", "load_error", "message"])
    size_column = _pick_present_column(table_columns, ["size_bytes", "size", "table_size"])
    query_count_column = _pick_present_column(table_columns, ["query_count"])
    recovery_source_column = _pick_present_column(table_columns, ["recovery_source"])
    load_start_column = _pick_present_column(table_columns, ["load_start_timestamp", "load_start_time", "load_start"])
    load_end_column = _pick_present_column(table_columns, ["load_end_timestamp", "load_end_time", "load_end"])

    if not (table_id_column and id_column):
        return report

    select_parts = []
    if schema_column:
        select_parts.append("i.{} as schema_name".format(_quote_identifier(schema_column)))
    if name_column:
        select_parts.append("i.{} as table_name".format(_quote_identifier(name_column)))
    if full_name_column:
        select_parts.append("i.{} as full_name".format(_quote_identifier(full_name_column)))
    if progress_column:
        select_parts.append("t.{} as load_progress".format(_quote_identifier(progress_column)))
    if status_column:
        select_parts.append("t.{} as status_text".format(_quote_identifier(status_column)))
    if error_column:
        select_parts.append("t.{} as error_text".format(_quote_identifier(error_column)))
    if size_column:
        select_parts.append("t.{} as size_bytes".format(_quote_identifier(size_column)))
    if query_count_column:
        select_parts.append("t.{} as query_count".format(_quote_identifier(query_count_column)))
    if recovery_source_column:
        select_parts.append("t.{} as recovery_source".format(_quote_identifier(recovery_source_column)))
    if load_start_column:
        select_parts.append("t.{} as load_start_timestamp".format(_quote_identifier(load_start_column)))
    if load_start_column and load_end_column:
        select_parts.append(
            "time_to_sec(timediff(t.{load_end}, t.{load_start})) as duration_in_sec".format(
                load_end=_quote_identifier(load_end_column),
                load_start=_quote_identifier(load_start_column),
            )
        )

    if not select_parts:
        return report

    rows = run_sql_dicts(
        """
        select {columns}
        from performance_schema.rpd_tables t
        inner join performance_schema.rpd_table_id i
            on i.{id_column} = t.{table_id_column}
        """.format(
            columns=", ".join(select_parts),
            id_column=_quote_identifier(id_column),
            table_id_column=_quote_identifier(table_id_column),
        ),
        include_database=False,
    )

    sorted_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("schema_name", "") or "").lower(),
            str(row.get("table_name", "") or row.get("full_name", "") or "").lower(),
        ),
    )
    for row in sorted_rows:
        schema_name = str(row.get("schema_name", "") or "").strip()
        table_name = str(row.get("table_name", "") or "").strip()
        if not schema_name or not table_name:
            derived_schema, derived_table = _split_heatwave_object_name(row.get("full_name", ""))
            schema_name = schema_name or derived_schema
            table_name = table_name or derived_table

        progress_value = _normalize_progress(row.get("load_progress"))
        status_text = str(row.get("status_text", "") or "").strip()
        error_text = str(row.get("error_text", "") or "").strip()
        row_class = _derive_heatwave_row_class(progress_value, status_text, error_text)
        report["rows"].append(
            [
                schema_name or "-",
                table_name or "-",
                "Loaded" if row_class == "loaded" else ("Error" if row_class == "error" else "Partial"),
                _format_progress(progress_value) if progress_value is not None else "-",
                status_text or "-",
                error_text or "-",
                row.get("size_bytes", "-") if row.get("size_bytes") not in (None, "") else "-",
                row.get("query_count", "-") if row.get("query_count") not in (None, "") else "-",
                str(row.get("recovery_source", "") or "-"),
                str(row.get("load_start_timestamp", "") or "-"),
                row.get("duration_in_sec", "-") if row.get("duration_in_sec") not in (None, "") else "-",
            ]
        )
        report["row_classes"].append(row_class)
        report["row_actions"].append(
            {
                "schema_name": schema_name,
                "table_name": table_name,
                "can_unload": bool(schema_name and table_name and not _is_system_database(schema_name)),
            }
        )
    return report


def fetch_heatwave_performance_queries():
    return run_sql_with_columns(
        """
        SELECT
            QUERY_ID,
            QUERY_TEXT,
            STR_TO_DATE(
                JSON_UNQUOTE(
                    JSON_EXTRACT(QEXEC_TEXT->>"$**.queryStartTime", '$[0]')
                ),
                '%Y-%m-%d %H:%i:%s.%f'
            ) AS QUERY_START,
            STR_TO_DATE(
                JSON_UNQUOTE(
                    JSON_EXTRACT(QEXEC_TEXT->>"$**.qexecStartTime", '$[0]')
                ),
                '%Y-%m-%d %H:%i:%s.%f'
            ) AS RPD_START,
            JSON_EXTRACT(QEXEC_TEXT->>"$**.timeBetweenMakePushedJoinAndRpdExecMsec", '$[0]')
            AS QUEUE_WAIT,
            STR_TO_DATE(
                JSON_UNQUOTE(
                    JSON_EXTRACT(QEXEC_TEXT->>"$**.queryEndTime", '$[0]')
                ),
                '%Y-%m-%d %H:%i:%s.%f'
            ) AS QUERY_END,
            JSON_EXTRACT(QEXEC_TEXT->>"$**.changePropagationSync.msec", '$[0]') AS ChangePropagation,
            JSON_EXTRACT(QEXEC_TEXT->>"$**.totalQueryTimeBreakdown.waitTime", '$[0]') AS Ttl_wait_time,
            JSON_EXTRACT(QEXEC_TEXT->>"$**.totalQueryTimeBreakdown.executionTime", '$[0]') AS Ttl_exec_time,
            JSON_EXTRACT(QEXEC_TEXT->>"$**.totalQueryTimeBreakdown.optimizationTime", '$[0]') AS Ttl_opt_time,
            JSON_EXTRACT(QEXEC_TEXT->>"$**.rpdExec.msec", '$[0]') AS RPD_EXEC,
            JSON_EXTRACT(QEXEC_TEXT->>"$**.getResults.msec", '$[0]') AS GET_RESULT,
            JSON_EXTRACT(QEXEC_TEXT->>"$**.sessionId", '$[0]') AS CONNECTION_ID,
            JSON_EXTRACT(QEXEC_TEXT->>"$**.qkrnActualRows[*].actRows", "$[0]") AS actRows
        FROM performance_schema.rpd_query_stats
        WHERE query_text not like 'ML_%'
        """,
        include_database=False,
    )


def fetch_heatwave_ml_queries(current_ml_connection_only=False):
    current_connection_filter = ""
    if current_ml_connection_only:
        current_connection_filter = """
          and connection_id = (
              select id
              from performance_schema.processlist
              where info like 'SET rapid_ml_operation%'
          )
        """
    return run_sql_with_columns(
        """
        select
            QEXEC_TEXT->>"$.startTime" as startTime,
            query_text,
            QEXEC_TEXT->>"$.status" as status,
            QEXEC_TEXT->>"$.totalRunTime" as totalRunTime,
            QEXEC_TEXT->>"$.details.operation" as operation,
            QEXEC_TEXT->>"$.completionPercentage" as completionPercentage,
            JSON_LENGTH(QEXEC_TEXT->>"$.progressItems") as progressItemsCount,
            JSON_LENGTH(QEXEC_TEXT->>"$.completedSteps") as completedSteps,
            JSON_EXTRACT(QEXEC_TEXT, concat("$.progressItems[", JSON_LENGTH(QEXEC_TEXT->>"$.completedSteps"), "]")) as progressItem,
            JSON_EXTRACT(
                JSON_EXTRACT(QEXEC_TEXT, concat("$.progressItems[", JSON_LENGTH(QEXEC_TEXT->>"$.completedSteps"), "]")),
                "$.type"
            ) as progressType,
            query_id,
            connection_id
        from performance_schema.rpd_query_stats
        where query_text like 'ML_%'
        {current_connection_filter}
        order by startTime desc
        """.format(current_connection_filter=current_connection_filter),
        include_database=False,
    )


def fetch_heatwave_ml_current_running_detail():
    return run_sql_with_columns(
        """
        select
            QEXEC_TEXT->>"$.startTime" as startTime,
            QEXEC_TEXT->>"$.status" as status,
            QEXEC_TEXT->>"$.completionPercentage" as completionPercentage,
            JSON_LENGTH(QEXEC_TEXT->>"$.progressItems") as progressItemsCount,
            JSON_LENGTH(QEXEC_TEXT->>"$.completedSteps") as completedSteps,
            JSON_EXTRACT(QEXEC_TEXT, concat("$.progressItems[", JSON_LENGTH(QEXEC_TEXT->>"$.completedSteps"), "]")) as progressItem,
            JSON_EXTRACT(
                JSON_EXTRACT(QEXEC_TEXT, concat("$.progressItems[", JSON_LENGTH(QEXEC_TEXT->>"$.completedSteps"), "]")),
                "$.type"
            ) as progressType,
            QEXEC_TEXT
        from performance_schema.rpd_query_stats
        where connection_id = (
            select id
            from performance_schema.processlist
            where info like 'SET rapid_ml_operation%'
        )
          and query_id = (
            select max(query_id)
            from performance_schema.rpd_query_stats
            where connection_id = (
                select id
                from performance_schema.processlist
                where info like 'SET rapid_ml_operation%'
            )
        )
        order by startTime desc
        limit 1
        """,
        include_database=False,
    )


def fetch_heatwave_table_load_recovery():
    return run_sql_with_columns(
        """
        select
            rpd_table_id.id,
            name,
            size_bytes,
            query_count,
            recovery_source,
            load_start_timestamp,
            time_to_sec(timediff(load_end_timestamp, load_start_timestamp)) as duration_in_sec
        from performance_schema.rpd_tables, performance_schema.rpd_table_id
        where rpd_tables.id = rpd_table_id.id
        order by size_bytes
        """,
        include_database=False,
    )


def fetch_table_definition(schema_name, table_name):
    rows = run_sql(
        """
        select
            column_name as column_name_value,
            column_type as column_type_value,
            data_type as data_type_value,
            is_nullable as is_nullable_value,
            column_key as column_key_value,
            column_default as column_default_value,
            extra as extra_value,
            ordinal_position as ordinal_position_value
        from information_schema.columns
        where table_schema = %s
          and table_name = %s
        order by ordinal_position
        """,
        (schema_name, table_name),
        include_database=False,
    )
    definition = []
    for row in rows:
        type_name, type_params = _split_mysql_data_type(row[1], row[2])
        definition.append(
            {
                "column_name": row[0],
                "column_type": row[1],
                "data_type": row[2],
                "type_name": type_name,
                "type_params": type_params,
                "is_nullable": row[3],
                "column_key": row[4],
                "column_default": row[5],
                "extra": row[6],
                "ordinal_position": row[7],
            }
        )
    return definition


def fetch_table_browse_page(schema_name, table_name, page_number=1, page_size=50):
    database_name = _validate_identifier(schema_name, "Database name")
    normalized_table_name = _validate_identifier(table_name, "Table name")
    if not _table_exists(database_name, normalized_table_name):
        raise ValueError(f"Table `{database_name}.{normalized_table_name}` was not found.")

    try:
        normalized_page = max(1, int(page_number or 1))
    except (TypeError, ValueError):
        normalized_page = 1
    offset = (normalized_page - 1) * page_size

    total_rows = run_sql(
        "select count(*) as row_count from {}.{}".format(
            _quote_identifier(database_name),
            _quote_identifier(normalized_table_name),
        ),
        include_database=False,
    )[0][0]
    page_count = max(1, (int(total_rows or 0) + page_size - 1) // page_size)
    normalized_page = min(normalized_page, page_count)
    offset = (normalized_page - 1) * page_size

    result = run_sql_with_columns(
        "select * from {}.{} limit %s offset %s".format(
            _quote_identifier(database_name),
            _quote_identifier(normalized_table_name),
        ),
        params=(page_size, offset),
        include_database=False,
    )
    return {
        "table_name": normalized_table_name,
        "columns": result["columns"],
        "rows": result["rows"],
        "page_number": normalized_page,
        "page_size": page_size,
        "page_count": page_count,
        "total_rows": int(total_rows or 0),
        "has_previous": normalized_page > 1,
        "has_next": normalized_page < page_count,
    }


def create_database(schema_name):
    database_name = _validate_identifier(schema_name, "Database name")
    exec_sql(
        f"create database {_quote_identifier(database_name)}",
        include_database=False,
    )
    return database_name


def drop_database(schema_name):
    database_name = _validate_identifier(schema_name, "Database name")
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be deleted.")
    exec_sql(
        f"drop database {_quote_identifier(database_name)}",
        include_database=False,
    )
    return database_name


def drop_table(schema_name, table_name):
    database_name = _validate_identifier(schema_name, "Database name")
    normalized_table_name = _validate_identifier(table_name, "Table name")
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be modified.")
    exec_sql(
        "drop table {}.{}".format(
            _quote_identifier(database_name),
            _quote_identifier(normalized_table_name),
        ),
        include_database=False,
    )
    return normalized_table_name


def collect_table_column_definitions(form):
    column_names = form.getlist("column_name")
    column_type_names = form.getlist("column_type_name")
    column_type_params = form.getlist("column_type_params")
    column_nullable = form.getlist("column_nullable")
    primary_indexes = set(form.getlist("column_primary"))
    columns = []
    max_length = max(len(column_names), len(column_type_names), len(column_type_params), len(column_nullable), 1)
    seen_names = set()

    for index, raw_name, raw_type_name, raw_type_params, raw_nullable in zip_longest(
        range(max_length),
        column_names,
        column_type_names,
        column_type_params,
        column_nullable,
        fillvalue="",
    ):
        name = str(raw_name or "").strip()
        type_name = str(raw_type_name or "").strip()
        type_params = str(raw_type_params or "").strip()
        nullable_choice = str(raw_nullable or "yes").strip().lower()
        is_primary = str(index) in primary_indexes

        if not name and not type_name and not type_params:
            continue
        if not name or not type_name:
            raise ValueError("Each table column needs both a name and a MySQL data type.")

        normalized_name = _validate_identifier(name, "Column name")
        normalized_data_type = _build_mysql_data_type(type_name, type_params)
        lowered_name = normalized_name.lower()
        if lowered_name in seen_names:
            raise ValueError(f"Duplicate column name: {normalized_name}.")
        seen_names.add(lowered_name)

        columns.append(
            {
                "name": normalized_name,
                "data_type": normalized_data_type,
                "type_name": str(type_name).upper(),
                "type_params": type_params,
                "nullable": nullable_choice != "no" and not is_primary,
                "primary": is_primary,
            }
        )

    if not columns:
        raise ValueError("Add at least one column before creating the table.")
    return columns


def create_table(schema_name, table_name, columns, *, add_invisible_auto_pk=False):
    database_name = _validate_identifier(schema_name, "Database name")
    normalized_table_name = _validate_identifier(table_name, "Table name")
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be modified.")
    if not columns:
        raise ValueError("Add at least one column before creating the table.")

    column_sql = []
    primary_key_columns = []
    for column in columns:
        definition = "{} {}".format(
            _quote_identifier(_validate_identifier(column["name"], "Column name")),
            _normalize_mysql_data_type(column["data_type"]),
        )
        if column.get("primary") or not column.get("nullable", True):
            definition += " NOT NULL"
        column_sql.append(definition)
        if column.get("primary"):
            primary_key_columns.append(_quote_identifier(column["name"]))

    ddl_parts = list(column_sql)
    if primary_key_columns:
        ddl_parts.append("PRIMARY KEY ({})".format(", ".join(primary_key_columns)))
    elif add_invisible_auto_pk:
        if any(str(column.get("name", "")).strip().lower() == "my_row_id" for column in columns):
            raise ValueError("Column name `my_row_id` is reserved for the generated invisible primary key.")
        ddl_parts.insert(
            0,
            "{} BIGINT UNSIGNED NOT NULL AUTO_INCREMENT INVISIBLE".format(_quote_identifier("my_row_id")),
        )
        ddl_parts.append("PRIMARY KEY ({})".format(_quote_identifier("my_row_id")))

    exec_sql(
        "create table {}.{} (\n  {}\n)".format(
            _quote_identifier(database_name),
            _quote_identifier(normalized_table_name),
            ",\n  ".join(ddl_parts),
        ),
        include_database=False,
    )
    return normalized_table_name


def add_table_column(schema_name, table_name, column):
    database_name = _validate_identifier(schema_name, "Database name")
    normalized_table_name = _validate_identifier(table_name, "Table name")
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be modified.")

    column_name = _validate_identifier(column["name"], "Column name")
    column_type = _build_mysql_data_type(column["type_name"], column.get("type_params", ""))
    definition = "{} {}".format(_quote_identifier(column_name), column_type)
    if not column.get("nullable", True):
        definition += " NOT NULL"
    exec_sql(
        "alter table {}.{} add column {}".format(
            _quote_identifier(database_name),
            _quote_identifier(normalized_table_name),
            definition,
        ),
        include_database=False,
    )
    return column_name


def modify_table_column(schema_name, table_name, original_column_name, column):
    database_name = _validate_identifier(schema_name, "Database name")
    normalized_table_name = _validate_identifier(table_name, "Table name")
    old_column_name = _validate_identifier(original_column_name, "Original column name")
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be modified.")

    new_column_name = _validate_identifier(column["name"], "Column name")
    column_type = _build_mysql_data_type(column["type_name"], column.get("type_params", ""))
    definition = "{} {}".format(_quote_identifier(new_column_name), column_type)
    if not column.get("nullable", True):
        definition += " NOT NULL"
    exec_sql(
        "alter table {}.{} change column {} {}".format(
            _quote_identifier(database_name),
            _quote_identifier(normalized_table_name),
            _quote_identifier(old_column_name),
            definition,
        ),
        include_database=False,
    )
    return new_column_name


def load_table_to_heatwave(schema_name, table_name, secondary_engine="RAPID"):
    database_name = _validate_identifier(schema_name, "Database name")
    normalized_table_name = _validate_identifier(table_name, "Table name")
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be modified.")

    if str(secondary_engine or "").strip().upper() != "RAPID":
        exec_sql(
            "alter table {}.{} secondary_engine=rapid".format(
                _quote_identifier(database_name),
                _quote_identifier(normalized_table_name),
            ),
            include_database=False,
        )
    exec_sql(
        "alter table {}.{} secondary_load".format(
            _quote_identifier(database_name),
            _quote_identifier(normalized_table_name),
        ),
        include_database=False,
    )
    return normalized_table_name


def unload_table_from_heatwave(schema_name, table_name):
    database_name = _validate_identifier(schema_name, "Database name")
    normalized_table_name = _validate_identifier(table_name, "Table name")
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be modified.")

    exec_sql(
        "alter table {}.{} secondary_unload".format(
            _quote_identifier(database_name),
            _quote_identifier(normalized_table_name),
        ),
        include_database=False,
    )
    return normalized_table_name


def load_database_to_heatwave(schema_name):
    database_name = _validate_identifier(schema_name, "Database name")
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be modified.")
    datasets = run_sql_multi_resultsets(
        'call sys.heatwave_load(json_array("{}"), null)'.format(database_name),
        include_database=False,
    )
    return {"database_name": database_name, "datasets": datasets}


def unload_database_from_heatwave(schema_name):
    database_name = _validate_identifier(schema_name, "Database name")
    if _is_system_database(database_name):
        raise ValueError("System databases cannot be modified.")
    datasets = run_sql_multi_resultsets(
        'call sys.heatwave_unload(json_array("{}"), null)'.format(database_name),
        include_database=False,
    )
    return {"database_name": database_name, "datasets": datasets}


def _build_table_model(rows, columns, *, labels=None, formatters=None):
    labels = labels or {}
    formatters = formatters or {}
    return {
        "columns": [labels.get(column, column) for column in columns],
        "rows": [
            [formatters.get(column, lambda value: value if value not in (None, "") else "-")(row.get(column)) for column in columns]
            for row in rows
        ],
    }


def _derive_cluster_status(node_rows, status_column):
    if not node_rows:
        return "No HeatWave nodes detected"
    if not status_column:
        return "HeatWave metadata available"
    statuses = sorted(
        {
            str(row.get(status_column, "")).strip()
            for row in node_rows
            if str(row.get(status_column, "")).strip()
        },
        key=str.lower,
    )
    return ", ".join(statuses) if statuses else "HeatWave metadata available"


def _derive_heatwave_traffic_light(status_rows):
    values = [str(row.get("value", "")).strip().upper() for row in status_rows]
    healthy_values = {"ON", "ENABLED", "AVAILABLE", "IDLE", "ONLINE"}
    if values and all(value in healthy_values for value in values):
        return {"state": "loaded", "label": "GREEN"}
    if values and all(value == "OFF" for value in values):
        return {"state": "error", "label": "RED"}
    return {"state": "partial", "label": "YELLOW"}


def get_dashboard_server_info():
    profile = get_session_profile()
    info = {
        "connection_endpoint": "{host}:{port}".format(**profile),
        "default_database": profile["database"] or "-",
        "user": session.get("db_user", "") or "-",
        "server_version": "-",
        "uptime": "-",
        "database_rows": [],
        "summary": {
            "database_count": 0,
            "table_count": 0,
            "data_length": 0,
            "index_length": 0,
            "total_length": 0,
            "database_count_display": "0",
            "table_count_display": "0",
            "data_length_display": "0 B",
            "index_length_display": "0 B",
            "total_length_display": "0 B",
        },
        "heatwave": {
            "available": False,
            "status": "HeatWave metadata not detected",
            "node_count": 0,
            "global_statuses": [
                {"name": "rapid_cluster_status", "value": "-"},
                {"name": "rapid_ml_status", "value": "-"},
                {"name": "rapid_change_propagation_status", "value": "-"},
                {"name": "rapid_net_encryption_status", "value": "-"},
                {"name": "rapid_preload_stats_status", "value": "-"},
                {"name": "rapid_resize_status", "value": "-"},
                {"name": "rapid_service_status", "value": "-"},
            ],
            "traffic_light": {"state": "partial", "label": "YELLOW"},
            "fully_loaded_count": None,
            "partially_loaded_count": None,
            "nodes_table": {"columns": [], "rows": []},
            "loaded_tables": {"columns": [], "rows": []},
            "partial_tables": {"columns": [], "rows": []},
            "notes": [],
        },
        "errors": [],
    }

    version_rows = run_sql(
        """
        select
            @@version as server_version,
            cast(variable_value as unsigned) as uptime_seconds
        from performance_schema.global_status
        where variable_name = 'Uptime'
        """,
        include_database=False,
    )
    if version_rows:
        info["server_version"] = version_rows[0][0] or "-"
        info["uptime"] = _format_uptime(version_rows[0][1])
    else:
        fallback_rows = run_sql(
            """
            select
                @@version as server_version,
                cast(variable_value as unsigned) as uptime_seconds
            from information_schema.global_status
            where variable_name = 'Uptime'
            """,
            include_database=False,
        )
        if fallback_rows:
            info["server_version"] = fallback_rows[0][0] or "-"
            info["uptime"] = _format_uptime(fallback_rows[0][1])

    database_rows = run_sql(
        """
        select
            s.schema_name as database_name,
            coalesce(stats.table_count, 0) as table_count,
            coalesce(stats.data_length, 0) as data_length,
            coalesce(stats.index_length, 0) as index_length
        from information_schema.schemata s
        left join (
            select
                table_schema,
                count(*) as table_count,
                coalesce(sum(data_length), 0) as data_length,
                coalesce(sum(index_length), 0) as index_length
            from information_schema.tables
            where table_type = 'BASE TABLE'
            group by table_schema
        ) stats
            on stats.table_schema = s.schema_name
        order by s.schema_name
        """,
        include_database=False,
    )
    summary = info["summary"]
    for database_name, table_count, data_length, index_length in database_rows:
        total_length = int(data_length or 0) + int(index_length or 0)
        is_system = database_name in SYSTEM_DATABASES
        info["database_rows"].append(
            {
                "database_name": database_name,
                "table_count": int(table_count or 0),
                "data_length": int(data_length or 0),
                "index_length": int(index_length or 0),
                "total_length": total_length,
                "data_length_display": _format_bytes(data_length),
                "index_length_display": _format_bytes(index_length),
                "total_length_display": _format_bytes(total_length),
                "is_system": is_system,
            }
        )
        if not is_system:
            summary["database_count"] += 1
            summary["table_count"] += int(table_count or 0)
            summary["data_length"] += int(data_length or 0)
            summary["index_length"] += int(index_length or 0)
            summary["total_length"] += total_length

    summary["database_count_display"] = str(summary["database_count"])
    summary["table_count_display"] = str(summary["table_count"])
    summary["data_length_display"] = _format_bytes(summary["data_length"])
    summary["index_length_display"] = _format_bytes(summary["index_length"])
    summary["total_length_display"] = _format_bytes(summary["total_length"])

    heatwave = info["heatwave"]
    rapid_status_rows = run_sql(
        """
        select variable_name, variable_value
        from performance_schema.global_status
        where lower(variable_name) like '%rapid%status%'
        order by variable_name
        """,
        include_database=False,
    )
    if not rapid_status_rows:
        rapid_status_rows = run_sql(
            """
            select variable_name, variable_value
            from information_schema.global_status
            where lower(variable_name) like '%rapid%status%'
            order by variable_name
            """,
            include_database=False,
        )
    if rapid_status_rows:
        heatwave["global_statuses"] = [
            {"name": str(variable_name or "-"), "value": str(variable_value or "-")}
            for variable_name, variable_value in rapid_status_rows
        ]
    heatwave["traffic_light"] = _derive_heatwave_traffic_light(heatwave["global_statuses"])

    has_rpd_nodes = _table_exists("performance_schema", "rpd_nodes")
    has_rpd_tables = _table_exists("performance_schema", "rpd_tables")
    has_rpd_table_id = _table_exists("performance_schema", "rpd_table_id")
    heatwave["available"] = has_rpd_nodes or has_rpd_tables or has_rpd_table_id

    if not heatwave["available"]:
        heatwave["notes"].append("performance_schema.rpd_nodes, performance_schema.rpd_tables, and performance_schema.rpd_table_id are not available on this server.")
        return info

    if has_rpd_nodes:
        node_columns = _get_table_columns("performance_schema", "rpd_nodes")
        node_id_column = _pick_present_column(
            node_columns,
            ["node_name", "node_id", "rpd_node_id", "host_name", "host", "address", "ip_address", "id"],
        )
        node_status_column = _pick_present_column(
            node_columns,
            ["status", "node_status", "health", "availability", "state"],
        )
        memory_columns = _pick_memory_columns(node_columns)
        selected_node_columns = _unique([node_id_column, node_status_column] + memory_columns[:4])
        if not selected_node_columns:
            selected_node_columns = node_columns[: min(len(node_columns), 6)]
        node_query = "select {} from {}.{}".format(
            ", ".join(_quote_identifier(column) for column in selected_node_columns),
            _quote_identifier("performance_schema"),
            _quote_identifier("rpd_nodes"),
        )
        node_rows = run_sql_dicts(node_query, include_database=False)
        node_count_rows = run_sql(
            "select count(*) as node_count from performance_schema.rpd_nodes",
            include_database=False,
        )
        heatwave["node_count"] = int(node_count_rows[0][0]) if node_count_rows else len(node_rows)
        heatwave["status"] = _derive_cluster_status(node_rows, node_status_column)
        heatwave["nodes_table"] = _build_table_model(
            node_rows,
            selected_node_columns,
            labels={
                node_id_column: "Node",
                node_status_column: "Status",
            },
            formatters={column: _format_bytes for column in memory_columns},
        )
        if not memory_columns:
            heatwave["notes"].append("No memory-related columns were exposed by performance_schema.rpd_nodes.")
    else:
        heatwave["notes"].append("performance_schema.rpd_nodes is not available.")

    if has_rpd_tables:
        table_columns = _get_table_columns("performance_schema", "rpd_tables")
        schema_column = _pick_present_column(
            table_columns,
            ["table_schema", "schema_name", "database_name", "db_name"],
        )
        table_name_column = _pick_present_column(
            table_columns,
            ["table_name", "name"],
        )
        table_id_column = _pick_present_column(
            table_columns,
            ["rpd_table_id", "table_id", "id"],
        )
        progress_column = _pick_present_column(
            table_columns,
            ["load_progress", "loading_progress", "load_percentage", "pct_loaded", "progress"],
        )
        status_column = _pick_present_column(
            table_columns,
            ["load_status", "loading_status", "status", "state"],
        )
        type_column = _pick_present_column(
            table_columns,
            ["load_type", "loading_type", "type"],
        )
        recovery_time_column = _pick_present_column(
            table_columns,
            ["recovery_time", "last_recovery_time", "recovery_start_time", "load_start_time"],
        )
        duration_column = _pick_present_column(
            table_columns,
            ["recovery_duration", "load_duration", "duration", "elapsed_time"],
        )
        selected_table_columns = _unique(
            [
                schema_column,
                table_name_column,
                table_id_column,
                status_column,
                progress_column,
                type_column,
                recovery_time_column,
                duration_column,
            ]
        )
        if not selected_table_columns:
            selected_table_columns = table_columns[: min(len(table_columns), 8)]
        table_query = "select {} from {}.{}".format(
            ", ".join(_quote_identifier(column) for column in selected_table_columns),
            _quote_identifier("performance_schema"),
            _quote_identifier("rpd_tables"),
        )
        table_rows = run_sql_dicts(table_query, include_database=False)
        partial_rows = []
        fully_loaded_count = 0
        partially_loaded_count = 0
        for row in table_rows:
            progress_value = _normalize_progress(row.get(progress_column)) if progress_column else None
            if progress_value is not None and progress_value >= 100:
                fully_loaded_count += 1
            elif progress_value is not None and 0 < progress_value < 100:
                partially_loaded_count += 1
                partial_rows.append(row)
        heatwave["fully_loaded_count"] = fully_loaded_count if progress_column else None
        heatwave["partially_loaded_count"] = partially_loaded_count if progress_column else None
        partial_table_columns = _unique(
            [
                schema_column,
                table_name_column,
                table_id_column,
                progress_column,
                type_column,
                recovery_time_column,
                duration_column,
                status_column,
            ]
        )
        if partial_table_columns:
            heatwave["partial_tables"] = _build_table_model(
                partial_rows,
                partial_table_columns,
                labels={
                    schema_column: "Schema",
                    table_name_column: "Table",
                    table_id_column: "RPD Table ID",
                    progress_column: "Load Progress",
                    type_column: "Type",
                    recovery_time_column: "Recovery Time",
                    duration_column: "Duration",
                    status_column: "Status",
                },
                formatters={progress_column: _format_progress} if progress_column else {},
            )
        if not progress_column:
            heatwave["notes"].append("No load-progress column was exposed by performance_schema.rpd_tables.")
    else:
        heatwave["notes"].append("performance_schema.rpd_tables is not available.")

    if has_rpd_tables and has_rpd_table_id:
        table_id_columns = _get_table_columns("performance_schema", "rpd_table_id")
        id_id_column = _pick_present_column(table_id_columns, ["id"])
        id_name_column = _pick_present_column(table_id_columns, ["name"])
        id_schema_column = _pick_present_column(table_id_columns, ["schema_name"])
        id_table_name_column = _pick_present_column(table_id_columns, ["table_name"])

        loaded_table_selects = []
        loaded_labels = {}
        loaded_formatters = {}

        if table_id_column:
            loaded_table_selects.append(
                "t.{} as rpd_table_id".format(_quote_identifier(table_id_column))
            )
            loaded_labels["rpd_table_id"] = "RPD Table ID"
        elif id_id_column:
            loaded_table_selects.append(
                "i.{} as rpd_table_id".format(_quote_identifier(id_id_column))
            )
            loaded_labels["rpd_table_id"] = "RPD Table ID"

        if id_name_column:
            loaded_table_selects.append(
                "i.{} as table_name".format(_quote_identifier(id_name_column))
            )
            loaded_labels["table_name"] = "Table"
        else:
            if id_schema_column:
                loaded_table_selects.append(
                    "i.{} as schema_name".format(_quote_identifier(id_schema_column))
                )
                loaded_labels["schema_name"] = "Schema"
            if id_table_name_column:
                loaded_table_selects.append(
                    "i.{} as short_table_name".format(_quote_identifier(id_table_name_column))
                )
                loaded_labels["short_table_name"] = "Table"

        if status_column:
            loaded_table_selects.append(
                "t.{} as load_status".format(_quote_identifier(status_column))
            )
            loaded_labels["load_status"] = "Status"
        if progress_column:
            loaded_table_selects.append(
                "t.{} as load_progress".format(_quote_identifier(progress_column))
            )
            loaded_labels["load_progress"] = "Load Progress"
            loaded_formatters["load_progress"] = _format_progress
        if type_column:
            loaded_table_selects.append(
                "t.{} as load_type".format(_quote_identifier(type_column))
            )
            loaded_labels["load_type"] = "Type"
        if recovery_time_column:
            loaded_table_selects.append(
                "t.{} as recovery_time".format(_quote_identifier(recovery_time_column))
            )
            loaded_labels["recovery_time"] = "Recovery Time"
        if duration_column:
            loaded_table_selects.append(
                "t.{} as duration".format(_quote_identifier(duration_column))
            )
            loaded_labels["duration"] = "Duration"

        if loaded_table_selects and (table_id_column or id_id_column):
            loaded_tables_query = """
            select {columns}
            from performance_schema.rpd_tables t
            inner join performance_schema.rpd_table_id i
                on i.{id_column} = t.{table_id_column}
            order by i.{name_order}, t.{status_order}
            """.format(
                columns=", ".join(loaded_table_selects),
                id_column=_quote_identifier(id_id_column or "ID"),
                table_id_column=_quote_identifier(table_id_column or "ID"),
                name_order=_quote_identifier(id_name_column or id_table_name_column or id_id_column or "ID"),
                status_order=_quote_identifier(status_column or table_id_column or "ID"),
            )
            loaded_table_rows = run_sql_dicts(loaded_tables_query, include_database=False)
            loaded_table_keys = [item.split(" as ", 1)[1] for item in loaded_table_selects]
            heatwave["loaded_tables"] = _build_table_model(
                loaded_table_rows,
                loaded_table_keys,
                labels=loaded_labels,
                formatters=loaded_formatters,
            )
        else:
            heatwave["notes"].append("Unable to determine a stable join key between performance_schema.rpd_tables and performance_schema.rpd_table_id.")
    elif has_rpd_tables and not has_rpd_table_id:
        heatwave["notes"].append("performance_schema.rpd_table_id is not available, so loaded table names cannot be shown.")

    return info


def get_nlsql_models():
    return get_generation_models()


def get_vision_models():
    return get_generation_models()


def _mysql_quote(value):
    text = str(value or "")
    return "'" + text.replace("\\", "\\\\").replace("'", "\\'") + "'"


def airportdb_exists(*, autocommit=False):
    rows = run_sql(
        """
        select schema_name as schema_name_value
        from information_schema.schemata
        where schema_name = %s
        """,
        ("airportdb",),
        include_database=False,
        autocommit=autocommit,
    )
    return bool(rows)


def build_nav_groups():
    groups = []
    show_performance = False
    object_storage_ready = False
    if session.get("logged_in"):
        try:
            show_performance = airportdb_exists()
        except mysql.connector.Error:
            show_performance = False
        object_storage_ready = askme_setup_is_ready()

    for group in NAV_GROUPS:
        items = []
        for item in group["items"]:
            item_payload = dict(item)
            if item["endpoint"] in {"askme_genai_page", "heatwave_lh_external_page"} and not object_storage_ready:
                item_payload["disabled"] = True
                item_payload["disabled_reason"] = "Configure Admin > Setup ObjectStorage first."
            items.append(item_payload)
        if group["label"] == "HeatWave" and show_performance:
            items.append(
                {
                    "endpoint": "heatwave_performance_page",
                    "label": "HeatWave Performance",
                }
            )
        groups.append({"label": group["label"], "items": items})
    return groups


def call_nlsql(question, model_id, schemas):
    dblist = ", ".join('"{}"'.format(schema_name) for schema_name in schemas)
    options = '{{"execute": true, "model_id": "{llm}", "schemas": [{schemas}]}}'.format(
        llm=model_id,
        schemas=dblist,
    )
    return call_proc("sys.NL_SQL", [question, "", options])


def build_nlsql_call_text(question, model_id, schemas):
    dblist = ", ".join('"{}"'.format(schema_name) for schema_name in schemas)
    options = '{{"execute": true, "model_id": "{llm}", "schemas": [{schemas}]}}'.format(
        llm=model_id,
        schemas=dblist,
    )
    return (
        "CALL sys.NL_SQL(\n"
        "  {question},\n"
        "  @output,\n"
        "  {options}\n"
        ");"
    ).format(
        question=_mysql_quote(question),
        options=_mysql_quote(options),
    )


def build_nlsql_tables(result):
    tables = []
    for index, rows in enumerate(result.get("resultset", [])):
        raw_columns = result.get("columnset", [])
        columns = list(raw_columns[index]) if index < len(raw_columns) else []
        tables.append(
            {
                "columns": columns,
                "rows": [list(row) for row in rows],
            }
        )
    return tables


def _queue_db_admin_modal_result(title, datasets):
    session["db_admin_modal_result"] = {
        "title": str(title or "Procedure Result"),
        "datasets": list(datasets or []),
        "message": "" if datasets else "Procedure completed without a tabular result set.",
    }


def _pop_db_admin_modal_result():
    return session.pop("db_admin_modal_result", None)


def _build_csv_response(filename, columns, rows):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([_normalize_modal_cell(value) for value in row])
    return Response(
        buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": 'attachment; filename="{}"'.format(filename)},
    )


def _build_db_admin_download_payload(active_tab, selected_database, monitor_view="heatwave-performance-query", current_ml_connection_only=False):
    if active_tab == "db":
        inventory = fetch_database_inventory()
        return (
            "db-admin-databases.csv",
            ["Database", "Tables", "Protected"],
            [[row["database_name"], row["table_count"], "Yes" if row["is_system"] else "No"] for row in inventory],
        )
    if active_tab == "table":
        if not selected_database:
            raise ValueError("Choose a database before downloading the table list.")
        tables = fetch_tables_for_database(selected_database)
        return (
            "db-admin-tables-{}.csv".format(selected_database),
            ["Table", "Rows", "Columns", "Engines", "HeatWave"],
            [
                [
                    row["table_name"],
                    row["row_count"],
                    row["column_count"],
                    row["engine_display"],
                    row["heatwave"]["progress_display"] if row["heatwave"]["progress_value"] is not None else "-",
                ]
                for row in tables
            ],
        )
    if active_tab == "hw-tables":
        report = fetch_heatwave_tables_report()
        return ("db-admin-hw-tables.csv", report["columns"], report["rows"])
    if active_tab == "monitoring":
        selected_monitor_view = str(monitor_view or "heatwave-performance-query").strip().lower()
        if selected_monitor_view not in DB_ADMIN_MONITORING_VIEWS:
            selected_monitor_view = "heatwave-performance-query"
        active_tab = selected_monitor_view
    if active_tab == "heatwave-performance-query":
        report = fetch_heatwave_performance_queries()
        return ("db-admin-heatwave-performance-query.csv", report["columns"], report["rows"])
    if active_tab == "heatwave-ml-query":
        report = fetch_heatwave_ml_queries(current_ml_connection_only=current_ml_connection_only)
        return ("db-admin-heatwave-ml-query.csv", report["columns"], report["rows"])
    if active_tab == "hw-table-load-recovery":
        report = fetch_heatwave_table_load_recovery()
        return ("db-admin-hw-table-load-recovery.csv", report["columns"], report["rows"])
    raise ValueError("Unsupported DB Admin tab for download.")


def answer_query_on_image(question, model_id, image_base64):
    rows = run_sql(
        """
        select sys.ML_GENERATE(
            %s,
            JSON_OBJECT('model_id', %s, 'image', %s)
        ) as response_value
        """,
        (question, model_id, image_base64),
    )
    return rows[0][0] if rows else ""


def read_app_version():
    try:
        payload = json.loads(VERSION_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return "unknown"
    if not isinstance(payload, dict):
        return "unknown"
    version = str(payload.get("version", "")).strip()
    return version or "unknown"


def _version_compare_key(value):
    parts = re.findall(r"\d+|[A-Za-z]+", str(value or ""))
    key = []
    for part in parts:
        if part.isdigit():
            key.append((1, int(part)))
        else:
            key.append((0, part.lower()))
    return key


def is_newer_version(candidate, current):
    if not candidate or not current or current == "unknown":
        return False
    return _version_compare_key(candidate) > _version_compare_key(current)


def _default_repo_version_url():
    configured_url = os.environ.get("HEATWAVE_DEMO_VERSION_URL", "").strip()
    if configured_url:
        return configured_url
    try:
        remote_url = subprocess.check_output(
            ["git", "config", "--get", "remote.origin.url"],
            cwd=str(ROOT_DIR),
            text=True,
            stderr=subprocess.DEVNULL,
            timeout=2,
        ).strip()
    except (OSError, subprocess.SubprocessError):
        remote_url = "https://github.com/ivanxma/HeatWave_Demo.git"
    try:
        branch_name = subprocess.check_output(
            ["git", "rev-parse", "--abbrev-ref", "HEAD"],
            cwd=str(ROOT_DIR),
            text=True,
            stderr=subprocess.DEVNULL,
            timeout=2,
        ).strip()
    except (OSError, subprocess.SubprocessError):
        branch_name = "main"
    if not branch_name or branch_name == "HEAD":
        branch_name = "main"
    match = re.match(r"git@github\.com:([^/]+)/(.+?)(?:\.git)?$", remote_url)
    if match:
        owner, repo = match.groups()
        return f"https://raw.githubusercontent.com/{owner}/{repo}/{branch_name}/{VERSION_FILE.name}"
    match = re.match(r"https://github\.com/([^/]+)/(.+?)(?:\.git)?$", remote_url)
    if match:
        owner, repo = match.groups()
        return f"https://raw.githubusercontent.com/{owner}/{repo}/{branch_name}/{VERSION_FILE.name}"
    return ""


def _get_git_upstream_ref():
    try:
        upstream = subprocess.check_output(
            ["git", "rev-parse", "--abbrev-ref", "--symbolic-full-name", "@{u}"],
            cwd=str(ROOT_DIR),
            text=True,
            stderr=subprocess.DEVNULL,
            timeout=2,
        ).strip()
    except (OSError, subprocess.SubprocessError):
        upstream = ""
    if upstream and "/" in upstream:
        remote_name, branch_name = upstream.split("/", 1)
        return remote_name, branch_name, upstream

    try:
        branch_name = subprocess.check_output(
            ["git", "rev-parse", "--abbrev-ref", "HEAD"],
            cwd=str(ROOT_DIR),
            text=True,
            stderr=subprocess.DEVNULL,
            timeout=2,
        ).strip()
    except (OSError, subprocess.SubprocessError):
        branch_name = ""
    if not branch_name or branch_name == "HEAD":
        branch_name = "main"
    return "origin", branch_name, f"origin/{branch_name}"


def _read_repo_version_from_git():
    remote_name, branch_name, upstream_ref = _get_git_upstream_ref()
    errors = []
    try:
        subprocess.run(
            ["git", "fetch", "--quiet", remote_name, branch_name],
            cwd=str(ROOT_DIR),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
            text=True,
            check=True,
            timeout=VERSION_CHECK_GIT_TIMEOUT_SECONDS,
        )
    except (OSError, subprocess.SubprocessError) as error:
        errors.append(str(error))

    for ref_name in (f"{remote_name}/{branch_name}", upstream_ref):
        try:
            version_text = subprocess.check_output(
                ["git", "show", f"{ref_name}:{VERSION_FILE.name}"],
                cwd=str(ROOT_DIR),
                text=True,
                stderr=subprocess.DEVNULL,
                timeout=2,
            )
            payload = json.loads(version_text)
        except (OSError, subprocess.SubprocessError, json.JSONDecodeError) as error:
            errors.append(str(error))
            continue
        if isinstance(payload, dict):
            repo_version = str(payload.get("version", "")).strip()
            if repo_version:
                return repo_version, ""
    return "", "; ".join(error for error in errors if error)


def check_repo_version():
    current_version = read_app_version()
    result = {
        "current_version": current_version,
        "repo_version": "",
        "update_available": False,
        "error": "",
    }
    repo_url = _default_repo_version_url()
    if not repo_url:
        result["error"] = "Repository version URL is not configured."
        return result
    errors = []
    try:
        request_obj = urllib.request.Request(repo_url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(request_obj, timeout=VERSION_CHECK_TIMEOUT_SECONDS) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception as error:
        errors.append(f"raw URL check failed: {error}")
        repo_version = ""
    else:
        if isinstance(payload, dict):
            repo_version = str(payload.get("version", "")).strip()
        else:
            repo_version = ""
    if not repo_version:
        git_version, git_error = _read_repo_version_from_git()
        repo_version = git_version
        if git_error:
            errors.append(f"git version check failed: {git_error}")
    result["repo_version"] = repo_version
    result["update_available"] = is_newer_version(repo_version, current_version)
    if not repo_version and errors:
        result["error"] = "Unable to check repository version. " + " ".join(errors)
    elif errors:
        result["error"] = " ".join(errors)
    return result


def render_dashboard(template_name, **context):
    connection_timeout_settings = fetch_connection_timeout_settings()
    app_version_status = session.get("version_check") if isinstance(session.get("version_check"), dict) else {}
    return render_template(
        template_name,
        app_title=APP_TITLE,
        app_version=read_app_version(),
        app_version_status=app_version_status,
        nav_groups=build_nav_groups(),
        current_user=session.get("db_user", ""),
        current_profile_name=session.get("profile_name", ""),
        connection_summary=get_connection_summary(),
        connection_timeout_summary=get_connection_timeout_summary(connection_timeout_settings),
        connection_timeout_settings=connection_timeout_settings,
        logged_in=bool(session.get("logged_in")),
        current_endpoint=request.endpoint or "",
        current_table="",
        **context,
    )


def _utc_now_text():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _parse_update_timestamp(value):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _mark_stale_restart_completed(status):
    if status.get("state") != "restarting":
        return status
    restart_time = _parse_update_timestamp(status.get("restart_requested_at") or status.get("updated_at"))
    if not restart_time:
        return status
    if datetime.now(timezone.utc) - restart_time < timedelta(seconds=30):
        return status
    status["state"] = "completed"
    status["step"] = "Completed"
    status["message"] = status.get("completion_message") or "Repository refresh, setup, and service restart completed."
    status["finished_at"] = status.get("finished_at") or _utc_now_text()
    return status


def _read_update_status():
    default_status = {
        "state": "idle",
        "step": "-",
        "message": "No update has been started.",
        "started_at": "",
        "updated_at": "",
        "finished_at": "",
        "service_names": [],
        "can_start": True,
        "log_text": "",
    }
    if UPDATE_STATUS_FILE.exists():
        try:
            payload = json.loads(UPDATE_STATUS_FILE.read_text(encoding="utf-8"))
            if isinstance(payload, dict):
                default_status.update(payload)
        except (OSError, json.JSONDecodeError):
            default_status["message"] = "The update status file could not be read."
            default_status["state"] = "error"
    if UPDATE_LOG_FILE.exists():
        try:
            default_status["log_text"] = UPDATE_LOG_FILE.read_text(encoding="utf-8", errors="replace")[-20000:]
        except OSError:
            default_status["log_text"] = "The update log could not be read."
    default_status = _mark_stale_restart_completed(default_status)
    default_status["can_start"] = default_status.get("state") not in {"starting", "running", "restarting"}
    return default_status


def _write_update_status(payload):
    UPDATE_STATUS_FILE.parent.mkdir(parents=True, exist_ok=True)
    temp_file = UPDATE_STATUS_FILE.with_suffix(".tmp")
    temp_file.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    temp_file.replace(UPDATE_STATUS_FILE)


def _start_update_worker():
    status = _read_update_status()
    if not status.get("can_start"):
        return False, "An update job is already running."
    if not UPDATE_WORKER_FILE.exists():
        return False, "The update worker script is missing."

    UPDATE_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    UPDATE_LOG_FILE.write_text("", encoding="utf-8")
    _write_update_status(
        {
            "state": "starting",
            "step": "Queued",
            "message": "Starting the update worker.",
            "started_at": _utc_now_text(),
            "updated_at": _utc_now_text(),
            "finished_at": "",
            "service_names": [],
            "can_start": False,
        }
    )
    subprocess.Popen(
        [
            sys.executable,
            str(UPDATE_WORKER_FILE),
            "--repo-dir",
            str(ROOT_DIR),
            "--status-file",
            str(UPDATE_STATUS_FILE),
            "--log-file",
            str(UPDATE_LOG_FILE),
            "--service-pid",
            str(os.getpid()),
        ],
        cwd=str(ROOT_DIR),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
        close_fds=True,
    )
    return True, "Update worker started."


@app.route("/admin/update", methods=["GET", "POST"])
@login_required
def update_heatwave_demo():
    if request.method == "POST":
        started, message = _start_update_worker()
        flash(message, "success" if started else "warning")
        return redirect(url_for("update_heatwave_demo"))
    return render_dashboard(
        "update_heatwave_demo.html",
        page_title="Update HeatWave_Demo",
        update_status=_read_update_status(),
        version_status=session.get("version_check") if isinstance(session.get("version_check"), dict) else {},
        status_url=url_for("update_heatwave_demo_status"),
        start_button_label="Update HeatWave_Demo",
    )


@app.route("/admin/update/status")
@login_required
def update_heatwave_demo_status():
    return jsonify(_read_update_status())


def redirect_for_profile_update(profile_name=""):
    return_to = request.form.get("return_to", "").strip()
    if return_to == "connection_profile" and session.get("logged_in"):
        return redirect(url_for("connection_profile", profile=profile_name))
    return redirect(url_for("login", profile=profile_name))


def get_heatwave_performance_table_counts():
    rows = run_sql_with_columns(
        """
        select 'booking' as table_name, count(*) as row_count from airportdb.booking
        union all
        select 'flight' as table_name, count(*) as row_count from airportdb.flight
        union all
        select 'airline' as table_name, count(*) as row_count from airportdb.airline
        union all
        select 'airport_geo' as table_name, count(*) as row_count from airportdb.airport_geo
        """,
        autocommit=True,
    )
    return [
        {"table_name": row[0], "row_count": row[1]}
        for row in rows["rows"]
    ]


def get_session_autocommit_value():
    rows = run_sql(
        "select @@session.autocommit as autocommit_value",
        autocommit=True,
    )
    return rows[0][0] if rows else None


def explain_heatwave_performance_query(sql_text):
    raw_plan = run_sql_with_columns("EXPLAIN " + sql_text, autocommit=True)
    formatted_rows = []
    for row in raw_plan["rows"]:
        formatted_row = []
        for value in row:
            cell_value = value
            is_multiline = False
            if isinstance(value, str):
                try:
                    parsed_json = json.loads(value)
                except json.JSONDecodeError:
                    parsed_json = None
                if isinstance(parsed_json, (dict, list)):
                    cell_value = json.dumps(parsed_json, indent=2)
                    is_multiline = True
            formatted_row.append({"value": cell_value, "is_multiline": is_multiline})
        formatted_rows.append(formatted_row)
    return {
        "columns": raw_plan["columns"],
        "rows": formatted_rows,
    }


def execute_heatwave_performance_query(sql_text):
    cnx = None
    cursor = None
    try:
        cnx = mysql_connection(autocommit=True)
        cursor = cnx.cursor()
        cursor.execute("select @@session.autocommit as autocommit_value")
        autocommit_value = cursor.fetchone()[0]
        started_at = datetime.now(timezone.utc)
        started_counter = time.perf_counter()
        cursor.execute(sql_text)
        rows = [list(row) for row in cursor.fetchall()]
        finished_counter = time.perf_counter()
        finished_at = datetime.now(timezone.utc)
        elapsed_seconds = (finished_at - started_at).total_seconds()
        return {
            "columns": list(cursor.column_names or ()),
            "rows": rows,
        }, {
            "started_at": started_at,
            "finished_at": finished_at,
            "elapsed_seconds": elapsed_seconds,
            "elapsed_perf_seconds": finished_counter - started_counter,
        }, autocommit_value
    finally:
        close_mysql_connection(cnx)


class DeferredTLSWSGIRequestHandler(WSGIRequestHandler):
    timeout = None

    def handle(self):
        try:
            if self.server.ssl_context is not None and isinstance(self.connection, ssl.SSLSocket):
                handshake_timeout = getattr(self.server, "ssl_handshake_timeout", None)
                if handshake_timeout is not None:
                    self.connection.settimeout(handshake_timeout)
                self.connection.do_handshake()
                self.connection.settimeout(getattr(self.server, "request_socket_timeout", None))
            BaseHTTPRequestHandler.handle(self)
        except (ConnectionError, socket.timeout) as error:
            self.connection_dropped(error)
        except Exception as error:
            if self.server.ssl_context is not None and is_ssl_error(error):
                self.log_error("SSL error occurred: %s", error)
            else:
                raise


class DeferredTLSThreadedWSGIServer(ThreadedWSGIServer):
    def __init__(
        self,
        host,
        port,
        app,
        handler=None,
        passthrough_errors=False,
        ssl_context=None,
        fd=None,
        *,
        ssl_handshake_timeout=DEFAULT_TLS_HANDSHAKE_TIMEOUT_SECONDS,
        request_socket_timeout=DEFAULT_REQUEST_SOCKET_TIMEOUT_SECONDS,
    ):
        resolved_ssl_context = None
        if ssl_context is not None:
            if isinstance(ssl_context, tuple):
                resolved_ssl_context = load_ssl_context(*ssl_context)
            else:
                resolved_ssl_context = ssl_context
        self.ssl_handshake_timeout = ssl_handshake_timeout
        self.request_socket_timeout = request_socket_timeout
        super().__init__(
            host,
            port,
            app,
            handler=handler or DeferredTLSWSGIRequestHandler,
            passthrough_errors=passthrough_errors,
            ssl_context=None,
            fd=fd,
        )
        if resolved_ssl_context is not None:
            self.socket = resolved_ssl_context.wrap_socket(
                self.socket,
                server_side=True,
                do_handshake_on_connect=False,
            )
            self.ssl_context = resolved_ssl_context


import pages.auth  # noqa: F401
import pages.askme_genai  # noqa: F401
import pages.connection_profile  # noqa: F401
import pages.db_admin  # noqa: F401
import pages.heatwave_genai  # noqa: F401
import pages.heatwave_performance  # noqa: F401
import pages.heatwave_lh_external  # noqa: F401
import pages.heatwave_ml  # noqa: F401
import pages.home  # noqa: F401
import pages.import_page  # noqa: F401
import pages.nlsql  # noqa: F401
import pages.setup_askme  # noqa: F401
import pages.setup_configdb  # noqa: F401
import pages.vision  # noqa: F401


if __name__ == "__main__":
    host = os.environ.get("APP_ADDRESS", "0.0.0.0")
    port = _normalized_port(os.environ.get("APP_PORT", "443"))
    cert_file = os.environ.get("APP_SSL_CERT_FILE", "")
    key_file = os.environ.get("APP_SSL_KEY_FILE", "")
    ssl_context = (cert_file, key_file) if cert_file and key_file else None
    if ssl_context is None:
        app.run(host=host, port=port, debug=False)
    else:
        ssl_handshake_timeout = (
            _normalized_optional_timeout(os.environ.get("APP_SSL_HANDSHAKE_TIMEOUT"))
            or DEFAULT_TLS_HANDSHAKE_TIMEOUT_SECONDS
        )
        request_socket_timeout = (
            _normalized_optional_timeout(os.environ.get("APP_REQUEST_SOCKET_TIMEOUT"))
            or DEFAULT_REQUEST_SOCKET_TIMEOUT_SECONDS
        )
        server = DeferredTLSThreadedWSGIServer(
            host,
            port,
            app,
            ssl_context=ssl_context,
            ssl_handshake_timeout=ssl_handshake_timeout,
            request_socket_timeout=request_socket_timeout,
        )
        server.serve_forever()
